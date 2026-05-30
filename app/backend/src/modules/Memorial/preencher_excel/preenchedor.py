import json
import openpyxl
import shutil
from pathlib import Path
from openpyxl.utils import get_column_letter

#Esse arquivo tem por objetivo carregar os dados do levantamento de campo (em formato JSON) e preencher o template de memorial descritivo (Excel) com esses dados, organizando-os nas abas e células corretas conforme o layout do template.
#Ele é projetado para ser executado como um script independente, mas pode ser adaptado para ser chamado a partir de outras partes do sistema conforme necessário.
#Funciona da seguinte forma:
#1. Carrega os dados do levantamento de campo a partir de um arquivo JSON (confira o arquivo com dados ficticios em `dados_levantamento_teste.json` para entender a estrutura esperada).
#2. Localiza o arquivo de template `model_memorial.xlsx` (deve estar na mesma pasta ou em pastas próximas).
#O Arquivo possui funcoes de preecnhimento de tabelo para cada aba. Por exemplo a funcao `popular_levantamento_campo` é responsavel por preencher a aba "Levantamento Campo" do Excel, mapeando os dados do JSON para as células corretas, respeitando o layout do template.
#As funcoes estão baseadas nas classes do arquivo models_abas.


def carregar_dados_json(caminho_json):
    """Carrega dados do arquivo JSON"""
    with open(caminho_json, 'r', encoding='utf-8') as f:
        return json.load(f)


def _criar_sheet(workbook, name):
    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(title=name)


def set_cell(sheet, row, col, value):
    """Escreve o valor respeitando células mescladas: escreve na célula superior-esquerda do range mesclado."""
    for mr in sheet.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            sheet.cell(row=mr.min_row, column=mr.min_col, value=value)
            return
    sheet.cell(row=row, column=col, value=value)


def popular_levantamento_campo(sheet, items):
    # Escreve somente nas coordenadas fixas do template (Tabela A e B)

    tabela_a = items.get('tabela_a', [])
    print(f"[debug] Registros Tabela A: {len(tabela_a)}")
    # Implementação: preencher linhas fixas para Tabela A
    rows_a = list(range(8, 28))
    ra = 0
    for reg in tabela_a:
        ambiente = reg.get('ambiente')
        dims = reg.get('dimensoes', [])
        vaos = reg.get('vaos', [])
        alvs = reg.get('alvenarias_adicionais', [])
        max_lines = max(len(dims), len(vaos), len(alvs))
        for i in range(max_lines):
            if ra >= len(rows_a):
                break
            r = rows_a[ra]
            # Ambiente (col B)
            set_cell(sheet, r, 2, ambiente)
            # Dimensões -> E,F,G,H (cols 5,6,7,8)
            if i < len(dims):
                d = dims[i]
                set_cell(sheet, r, 5, d.get('c'))
                set_cell(sheet, r, 6, d.get('l'))
                set_cell(sheet, r, 7, d.get('h'))
                set_cell(sheet, r, 8, d.get('e'))
            # Vãos -> J,K,L,M (cols 10,11,12,13)
            if i < len(vaos):
                v = vaos[i]
                set_cell(sheet, r, 10, v.get('tipo'))
                set_cell(sheet, r, 11, v.get('c'))
                set_cell(sheet, r, 12, v.get('h'))
                set_cell(sheet, r, 13, v.get('e'))
            # Alvenarias Adicionais -> O,P,Q,R (cols 15,16,17,18)
            if i < len(alvs):
                a = alvs[i]
                set_cell(sheet, r, 15, a.get('tipo'))
                set_cell(sheet, r, 16, a.get('c'))
                set_cell(sheet, r, 17, a.get('h'))
                set_cell(sheet, r, 18, a.get('e'))
            ra += 1

    # Tabela B: escrever nas linhas fixas 33..52 conforme mapeamento
    rows_b = list(range(33, 53))
    tabela_b = items.get('tabela_b', [])
    rb = 0
    print(f"[debug] Registros Tabela B: {len(tabela_b)}")
    for reg in tabela_b:
        pilares = reg.get('pilares', [])
        vigas = reg.get('vigas', [])
        lajes = reg.get('lajes', [])
        max_lines = max(len(pilares), len(vigas), len(lajes))
        for i in range(max_lines):
            if rb >= len(rows_b):
                break
            r = rows_b[rb]
            # Pilar: C->E (col5), L->F (6), h->G (7)
            if i < len(pilares):
                p = pilares[i]
                set_cell(sheet, r, 5, p.get('c'))
                set_cell(sheet, r, 6, p.get('l'))
                set_cell(sheet, r, 7, p.get('h'))
            # Viga: C->I (9), h->J (10), e->K (11)
            if i < len(vigas):
                v = vigas[i]
                set_cell(sheet, r, 9, v.get('c'))
                set_cell(sheet, r, 10, v.get('h'))
                set_cell(sheet, r, 11, v.get('e'))
            # Laje: C->M (13), L->N (14), e->O (15)
            if i < len(lajes):
                l = lajes[i]
                set_cell(sheet, r, 13, l.get('c'))
                set_cell(sheet, r, 14, l.get('l'))
                set_cell(sheet, r, 15, l.get('e'))
            rb += 1

    print(f"[debug] Preenchimento completo nas áreas A e B")

    # --- Tabela C: linhas 58..77 ---
    rows_c = list(range(58, 78))
    tabela_c = items.get('tabela_c', [])
    print(f"[debug] Registros Tabela C: {len(tabela_c)}")
    rc = 0
    for reg in tabela_c:
        # campos: Quadros:[E], Conduites:[F], Tomadas[G], Interruptor[H], Luminárias:[I], Dutos:[J], Cabos:[K]
        # Acessorios: Tipo:[L], Un:[N]
        # Equipamentos: Tipo:[O], Un:[Q]
        for i in range(max(1, len(reg.get('itens', [1])))):
            if rc >= len(rows_c):
                break
            r = rows_c[rc]
            set_cell(sheet, r, 2, reg.get('ambiente'))
            set_cell(sheet, r, 5, reg.get('quadros'))
            set_cell(sheet, r, 6, reg.get('conduites'))
            set_cell(sheet, r, 7, reg.get('tomadas'))
            set_cell(sheet, r, 8, reg.get('interruptores'))
            set_cell(sheet, r, 9, reg.get('luminarias'))
            set_cell(sheet, r, 10, reg.get('dutos'))
            set_cell(sheet, r, 11, reg.get('cabos'))
            # acessorios
            set_cell(sheet, r, 12, reg.get('acessorios', {}).get('tipo'))
            set_cell(sheet, r, 14, reg.get('acessorios', {}).get('un'))
            # equipamentos
            set_cell(sheet, r, 15, reg.get('equipamentos', {}).get('tipo'))
            set_cell(sheet, r, 17, reg.get('equipamentos', {}).get('un'))
            rc += 1

    # --- Tabela D: linhas 84..103 ---
    rows_d = list(range(84, 104))
    tabela_d = items.get('tabela_d', [])
    print(f"[debug] Registros Tabela D: {len(tabela_d)}")
    rd = 0
    for reg in tabela_d:
        for i in range(max(1, len(reg.get('itens', [1])))):
            if rd >= len(rows_d):
                break
            r = rows_d[rd]
            set_cell(sheet, r, 2, reg.get('ambiente'))
            # Água Fria/Quente/Reúso
            set_cell(sheet, r, 5, reg.get('agua', {}).get('cavalete'))
            set_cell(sheet, r, 6, reg.get('agua', {}).get('reservatorio', {}).get('un'))
            set_cell(sheet, r, 7, reg.get('agua', {}).get('reservatorio', {}).get('l'))
            set_cell(sheet, r, 8, reg.get('agua', {}).get('registros'))
            set_cell(sheet, r, 9, reg.get('agua', {}).get('valvulas'))
            set_cell(sheet, r, 10, reg.get('agua', {}).get('torneiras'))
            set_cell(sheet, r, 11, reg.get('agua', {}).get('dutos'))
            # Água Pluvial
            set_cell(sheet, r, 12, reg.get('pluvial', {}).get('calhas', {}).get('tipo'))
            set_cell(sheet, r, 13, reg.get('pluvial', {}).get('calhas', {}).get('m'))
            set_cell(sheet, r, 14, reg.get('pluvial', {}).get('dutos'))
            set_cell(sheet, r, 15, reg.get('pluvial', {}).get('caixas'))
            # Esgoto
            set_cell(sheet, r, 16, reg.get('esgoto', {}).get('drenos'))
            set_cell(sheet, r, 17, reg.get('esgoto', {}).get('dutos'))
            set_cell(sheet, r, 18, reg.get('esgoto', {}).get('caixas'))
            rd += 1

    # --- Tabela E: linhas 110..129 ---
    rows_e = list(range(110, 130))
    tabela_e = items.get('tabela_e', [])
    print(f"[debug] Registros Tabela E: {len(tabela_e)}")
    re = 0
    for reg in tabela_e:
        for i in range(max(1, len(reg.get('itens', [1])))):
            if re >= len(rows_e):
                break
            r = rows_e[re]
            set_cell(sheet, r, 2, reg.get('ambiente'))
            set_cell(sheet, r, 5, reg.get('rede', {}).get('quadros'))
            set_cell(sheet, r, 6, reg.get('rede', {}).get('condulete'))
            set_cell(sheet, r, 7, reg.get('rede', {}).get('tomadas'))
            set_cell(sheet, r, 8, reg.get('rede', {}).get('dutos', {}).get('tipo'))
            set_cell(sheet, r, 9, reg.get('rede', {}).get('dutos', {}).get('m'))
            set_cell(sheet, r, 10, reg.get('rede', {}).get('cabos'))
            # SPDA
            set_cell(sheet, r, 11, reg.get('spda', {}).get('captacao'))
            set_cell(sheet, r, 12, reg.get('spda', {}).get('condulete'))
            set_cell(sheet, r, 13, reg.get('spda', {}).get('aterramento'))
            set_cell(sheet, r, 14, reg.get('spda', {}).get('dutos', {}).get('tipo'))
            set_cell(sheet, r, 15, reg.get('spda', {}).get('dutos', {}).get('m'))
            set_cell(sheet, r, 16, reg.get('spda', {}).get('cabos'))
            re += 1

    # --- Tabela F: linhas 136..155 ---
    rows_f = list(range(136, 156))
    tabela_f = items.get('tabela_f', [])
    print(f"[debug] Registros Tabela F: {len(tabela_f)}")
    rf = 0
    for reg in tabela_f:
        for i in range(max(1, len(reg.get('itens', [1])))):
            if rf >= len(rows_f):
                break
            r = rows_f[rf]
            set_cell(sheet, r, 2, reg.get('ambiente'))
            # Contra Incêndio
            set_cell(sheet, r, 5, reg.get('contra_incendio', {}).get('reservatorio', {}).get('tipo'))
            set_cell(sheet, r, 6, reg.get('contra_incendio', {}).get('reservatorio', {}).get('un'))
            set_cell(sheet, r, 7, reg.get('contra_incendio', {}).get('registros'))
            set_cell(sheet, r, 8, reg.get('contra_incendio', {}).get('valvulas'))
            set_cell(sheet, r, 9, reg.get('contra_incendio', {}).get('dutos', {}).get('tipo'))
            set_cell(sheet, r, 10, reg.get('contra_incendio', {}).get('dutos', {}).get('m'))
            set_cell(sheet, r, 11, reg.get('contra_incendio', {}).get('hidrantes'))
            # Instalações Pressurizadas
            set_cell(sheet, r, 12, reg.get('pressurizadas', {}).get('reservatorio', {}).get('tipo'))
            set_cell(sheet, r, 13, reg.get('pressurizadas', {}).get('reservatorio', {}).get('un'))
            set_cell(sheet, r, 14, reg.get('pressurizadas', {}).get('registros'))
            set_cell(sheet, r, 15, reg.get('pressurizadas', {}).get('valvulas'))
            set_cell(sheet, r, 16, reg.get('pressurizadas', {}).get('dutos', {}).get('tipo'))
            set_cell(sheet, r, 17, reg.get('pressurizadas', {}).get('dutos', {}).get('m'))
            set_cell(sheet, r, 18, reg.get('pressurizadas', {}).get('hidrantes'))
            rf += 1

    # --- Tabela G: linhas 162..171 ---
    rows_g = list(range(162, 172))
    tabela_g = items.get('tabela_g', [])
    print(f"[debug] Registros Tabela G: {len(tabela_g)}")
    rg = 0
    for reg in tabela_g:
        for i in range(max(1, len(reg.get('itens', [1])))):
            if rg >= len(rows_g):
                break
            r = rows_g[rg]
            set_cell(sheet, r, 2, reg.get('ambiente'))
            # Telhados
            set_cell(sheet, r, 5, reg.get('telhados', {}).get('c'))
            set_cell(sheet, r, 6, reg.get('telhados', {}).get('l'))
            set_cell(sheet, r, 7, reg.get('telhados', {}).get('h'))
            # Estrutura
            # Tipo em H162:J171 -> mapear para col 8 (H)
            set_cell(sheet, r, 8, reg.get('estrutura', {}).get('tipo'))
            set_cell(sheet, r, 9, reg.get('estrutura', {}).get('l'))
            set_cell(sheet, r, 10, reg.get('estrutura', {}).get('c'))
            set_cell(sheet, r, 11, reg.get('estrutura', {}).get('e'))
            # Telhamento
            set_cell(sheet, r, 13, reg.get('telhamento', {}).get('tipo'))
            set_cell(sheet, r, 14, reg.get('telhamento', {}).get('l'))
            set_cell(sheet, r, 15, reg.get('telhamento', {}).get('c'))
            set_cell(sheet, r, 16, reg.get('telhamento', {}).get('e'))
            rg += 1


def popular_servicos_preliminares(sheet, lista):
    # Suporta duas formas de entrada: 1) `lista` é um dict com chaves 'interdicoes','remocoes','demolicoes'
    # ou 2) `lista` já é a lista de interdições (compatibilidade retroativa).
    if isinstance(lista, dict):
        interdicoes = lista.get('interdicoes', [])
        remocoes = lista.get('remocoes', [])
        demolicoes = lista.get('demolicoes', [])
    else:
        interdicoes = lista
        remocoes = []
        demolicoes = []

    # --- Tabela Interdições: linhas 38..41 ---
    rows_int = list(range(38, 42))
    for idx, it in enumerate(interdicoes):
        if idx >= len(rows_int):
            break
        r = rows_int[idx]
        # Fase -> B (2), Ambiente -> C (3), Local -> E (5) (pode ser célula mesclada)
        set_cell(sheet, r, 2, it.get('fase'))
        set_cell(sheet, r, 3, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('local'))

    # --- Tabela Remoções: linhas 49..68 ---
    rows_rem = list(range(49, 69))
    for idx, reg in enumerate(remocoes):
        if idx >= len(rows_rem):
            break
        r = rows_rem[idx]
        # Ambiente (col B)
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Elétrica / SPDA / Rede
        set_cell(sheet, r, 5, reg.get('eletrica', {}).get('condulete'))
        set_cell(sheet, r, 6, reg.get('eletrica', {}).get('tomadas'))
        set_cell(sheet, r, 7, reg.get('eletrica', {}).get('interruptores'))
        set_cell(sheet, r, 8, reg.get('eletrica', {}).get('luminarias'))
        set_cell(sheet, r, 9, reg.get('eletrica', {}).get('dutos'))
        set_cell(sheet, r, 10, reg.get('eletrica', {}).get('cabos'))
        set_cell(sheet, r, 11, reg.get('eletrica', {}).get('captacao'))
        set_cell(sheet, r, 12, reg.get('eletrica', {}).get('aterramento'))
        set_cell(sheet, r, 13, reg.get('eletrica', {}).get('quadros'))
        set_cell(sheet, r, 14, reg.get('eletrica', {}).get('postes'))
        # Água / Pluvial / Esgoto / Pressurizadas / Contra Incêndio
        set_cell(sheet, r, 15, reg.get('agua', {}).get('cavalete'))
        set_cell(sheet, r, 16, reg.get('agua', {}).get('reservatorio'))
        set_cell(sheet, r, 17, reg.get('agua', {}).get('registros'))
        set_cell(sheet, r, 18, reg.get('agua', {}).get('valvulas'))
        set_cell(sheet, r, 19, reg.get('agua', {}).get('torneiras'))
        set_cell(sheet, r, 20, reg.get('agua', {}).get('dutos'))
        set_cell(sheet, r, 21, reg.get('agua', {}).get('calhas'))
        set_cell(sheet, r, 22, reg.get('agua', {}).get('caixas'))
        set_cell(sheet, r, 23, reg.get('agua', {}).get('drenos'))
        # Esquadrias
        set_cell(sheet, r, 24, reg.get('esquadrias', {}).get('portas'))
        set_cell(sheet, r, 25, reg.get('esquadrias', {}).get('janelas'))
        # Telhados
        set_cell(sheet, r, 26, reg.get('telhados', {}).get('telha'))
        set_cell(sheet, r, 27, reg.get('telhados', {}).get('trama'))
        set_cell(sheet, r, 28, reg.get('telhados', {}).get('tesoura'))
        # Equipamentos/Acessórios
        set_cell(sheet, r, 29, reg.get('equipamentos', {}).get('item'))
        set_cell(sheet, r, 31, reg.get('equipamentos', {}).get('qtd'))

    # --- Tabela Demolições: linhas 76..95 ---
    rows_dem = list(range(76, 96))
    for idx, reg in enumerate(demolicoes):
        if idx >= len(rows_dem):
            break
        r = rows_dem[idx]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        set_cell(sheet, r, 5, reg.get('piso'))
        set_cell(sheet, r, 6, reg.get('rodape'))
        set_cell(sheet, r, 7, reg.get('azulejo'))
        set_cell(sheet, r, 8, reg.get('forro'))
        # Alvenaria
        set_cell(sheet, r, 9, reg.get('alvenaria', {}).get('tipo'))
        set_cell(sheet, r, 10, reg.get('alvenaria', {}).get('v'))
        # Estrutura
        set_cell(sheet, r, 11, reg.get('estrutura', {}).get('fundacao'))
        set_cell(sheet, r, 12, reg.get('estrutura', {}).get('pilar'))
        set_cell(sheet, r, 13, reg.get('estrutura', {}).get('viga'))
        set_cell(sheet, r, 14, reg.get('estrutura', {}).get('laje'))


def popular_movimento_solo(sheet, lista):
    # Aceita `lista` como dict com chaves por subtabela ou lista (compatibilidade)
    if isinstance(lista, dict):
        escavacoes = lista.get('escavacoes', [])
        aterros = lista.get('aterros', [])
        enrocamentos = lista.get('enrocamentos', [])
        contencoes = lista.get('contencoes', [])
        taludamentos = lista.get('taludamentos', [])
        nivelamentos = lista.get('nivelamentos', [])
    else:
        # Se for uma lista, assumimos que seja lista de escavações
        escavacoes = lista
        aterros = enrocamentos = contencoes = taludamentos = nivelamentos = []

    # --- Escavações: linhas 17..36 ---
    rows_esc = list(range(17, 37))
    for idx, it in enumerate(escavacoes):
        if idx >= len(rows_esc):
            break
        r = rows_esc[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('tipo'))     # E
        set_cell(sheet, r, 6, it.get('i_pct'))   # F
        set_cell(sheet, r, 7, it.get('l_m'))     # G
        set_cell(sheet, r, 8, it.get('c_m'))     # H
        set_cell(sheet, r, 9, it.get('h_m'))     # I
        set_cell(sheet, r, 10, it.get('lastro_m'))# J
        set_cell(sheet, r, 11, it.get('area_m2'))# K
        set_cell(sheet, r, 12, it.get('volume_m3'))# L

    # --- Aterros e Reaterros: linhas 43..62 ---
    rows_ater = list(range(43, 63))
    for idx, it in enumerate(aterros):
        if idx >= len(rows_ater):
            break
        r = rows_ater[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('i_pct'))
        set_cell(sheet, r, 6, it.get('l_m'))
        set_cell(sheet, r, 7, it.get('c_m'))
        set_cell(sheet, r, 8, it.get('h_m'))
        set_cell(sheet, r, 9, it.get('area_m2'))
        set_cell(sheet, r, 10, it.get('volume_m3'))

    # --- Enrocamentos: linhas 69..88 ---
    rows_enr = list(range(69, 89))
    for idx, it in enumerate(enrocamentos):
        if idx >= len(rows_enr):
            break
        r = rows_enr[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('i_pct'))
        set_cell(sheet, r, 6, it.get('l_m'))
        set_cell(sheet, r, 7, it.get('c_m'))
        set_cell(sheet, r, 8, it.get('h_m'))
        set_cell(sheet, r, 9, it.get('area_m2'))
        set_cell(sheet, r, 10, it.get('volume_m3'))

    # --- Contenções: linhas 95..114 ---
    rows_con = list(range(95, 115))
    for idx, it in enumerate(contencoes):
        if idx >= len(rows_con):
            break
        r = rows_con[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('i_pct'))
        set_cell(sheet, r, 6, it.get('l_m'))
        set_cell(sheet, r, 7, it.get('c_m'))
        set_cell(sheet, r, 8, it.get('h_m'))
        set_cell(sheet, r, 9, it.get('area_m2'))
        set_cell(sheet, r, 10, it.get('volume_m3'))

    # --- Taludamentos: linhas 121..141 ---
    rows_tal = list(range(121, 142))
    for idx, it in enumerate(taludamentos):
        if idx >= len(rows_tal):
            break
        r = rows_tal[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('i_pct'))
        set_cell(sheet, r, 6, it.get('l_m'))
        set_cell(sheet, r, 7, it.get('c_m'))
        set_cell(sheet, r, 8, it.get('h_m'))
        set_cell(sheet, r, 9, it.get('area_m2'))
        set_cell(sheet, r, 10, it.get('volume_m3'))

    # --- Nivelamentos e Compactações: linhas 148..167 ---
    rows_niv = list(range(148, 168))
    for idx, it in enumerate(nivelamentos):
        if idx >= len(rows_niv):
            break
        r = rows_niv[idx]
        set_cell(sheet, r, 2, it.get('ambiente'))
        set_cell(sheet, r, 5, it.get('i_pct'))
        set_cell(sheet, r, 6, it.get('l_m'))
        set_cell(sheet, r, 7, it.get('c_m'))
        set_cell(sheet, r, 8, it.get('h_m'))
        set_cell(sheet, r, 9, it.get('area_m2'))
        set_cell(sheet, r, 10, it.get('volume_m3'))


def popular_estruturas(sheet, estruturas):
    """Preenche a aba Estruturas dividida em várias partes com ranges fixos.

    O parâmetro `estruturas` deve ser um dict contendo listas para cada parte.
    Procura chaves 'parte1'..'parte25' e cai em fallback genérico quando não existirem.
    """

    def get_nested(d, path, default=None):
        if not d or not path:
            return default
        cur = d
        for p in path.split('.'):
            if not isinstance(cur, dict):
                return default
            cur = cur.get(p, default)
        return cur

    def fill_range(start_row, end_row, data_list, mapping):
        rows = list(range(start_row, end_row + 1))
        for idx, item in enumerate(data_list):
            if idx >= len(rows):
                break
            r = rows[idx]
            for key, col in mapping.items():
                val = get_nested(item, key)
                set_cell(sheet, r, col, val)

    # Helper to get list for a part with fallbacks
    def part_list(name):
        return estruturas.get(name) or estruturas.get(name.replace('parte', 'parte_')) or estruturas.get(name.replace('parte', 'p')) or estruturas.get('itens', []) or estruturas.get('lista', []) or []

    # Parte 1: B12..B31 (rows 12..31)
    fill_range(12, 31, part_list('parte1'), {
        'ambiente': 2,
        'peca': 5,
        'secao.l': 6,
        'secao.h': 7,
        'c': 8,
        'lastro': 9,
        'concreto': 10,
        'ferragem': 11,
        'estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 2: rows 39..58
    fill_range(39, 58, part_list('parte2'), {
        'ambiente': 2,
        'peca': 5,
        'valores_por_peca.l': 6,
        'valores_por_peca.h': 7,
        'valores_por_peca.c': 8,
        'valores_totais.lastro': 9,
        'valores_totais.concreto': 10,
        'valores_totais.ferragem': 11,
        'valores_totais.estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 3: rows 66..85
    fill_range(66, 85, part_list('parte3'), {
        'ambiente': 2,
        'peca': 5,
        'valores_por_peca.l': 6,
        'valores_por_peca.h': 7,
        'valores_por_peca.c': 8,
        'valores_totais.lastro': 9,
        'valores_totais.concreto': 10,
        'valores_totais.ferragem': 11,
        'valores_totais.estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 4: rows 93..112
    fill_range(93, 112, part_list('parte4'), {
        'ambiente': 2,
        'peca': 5,
        'valores_por_peca.l': 6,
        'valores_por_peca.h': 7,
        'valores_por_peca.c': 8,
        'valores_totais.lastro': 9,
        'valores_totais.concreto': 10,
        'valores_totais.ferragem': 11,
        'valores_totais.estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 5: rows 120..139
    fill_range(120, 139, part_list('parte5'), {
        'ambiente': 2,
        'peca': 5,
        'valores_por_peca.l': 6,
        'valores_por_peca.h': 7,
        'valores_por_peca.c': 8,
        'valores_totais.lastro': 9,
        'valores_totais.concreto': 10,
        'valores_totais.ferragem': 11,
        'valores_totais.estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 6: rows 147..166
    fill_range(147, 166, part_list('parte6'), {
        'ambiente': 2,
        'peca': 5,
        'valores_por_peca.l': 6,
        'valores_por_peca.h': 7,
        'valores_por_peca.c': 8,
        'valores_totais.lastro': 9,
        'valores_totais.concreto': 10,
        'valores_totais.ferragem': 11,
        'valores_totais.estribo': 12,
        'forma_madeira.l': 13,
        'forma_madeira.c': 14,
    })

    # Parte 7: rows 176..195 (note: janela lança/o at col O -> 15)
    fill_range(176, 195, part_list('parte7'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 8: rows 203..222
    fill_range(203, 222, part_list('parte8'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 9: rows 230..249
    fill_range(230, 249, part_list('parte9'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 10: rows 257..276
    fill_range(257, 276, part_list('parte10'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 11: rows 284..303
    fill_range(284, 303, part_list('parte11'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 12: rows 311..330
    fill_range(311, 330, part_list('parte12'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 13: rows 338..357
    fill_range(338, 357, part_list('parte13'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 14: rows 375..394
    fill_range(375, 394, part_list('parte14'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 15: rows 402..421
    fill_range(402, 421, part_list('parte15'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 16: rows 429..448
    fill_range(429, 448, part_list('parte16'), {
        'ambiente': 2,
        'peca': 5,
        'l': 6,
        'h': 7,
        'c': 8,
        'concreto': 9,
        'ferragem': 10,
        'estribo': 11,
        'forma_madeira.c': 12,
        'forma_madeira.l': 13,
        'janela_lanca': 15,
    })

    # Parte 17..25: estruturas com perfil/ seção/ peso e equipamento de apoio (rows vary)
    part_defs = [
        ('parte17', 458, 477), ('parte18', 485, 504), ('parte19', 512, 531), ('parte20', 539, 558),
        ('parte21', 566, 585), ('parte22', 593, 612), ('parte23', 620, 639), ('parte24', 647, 666),
        ('parte25', 675, 691)
    ]
    for name, start, end in part_defs:
        fill_range(start, end, part_list(name), {
            'ambiente': 2,
            'peca': 5,
            'h': 6 if name not in ('parte17',) else 6,  # h usually column F
            'perfil': 7,
            'secao.l': 8,
            'secao.c': 9,
            'peso': 10,
            'elastometro': 11,
            'equipamento_apoi.h': 12,
            'equipamento_apoi.c': 13,
            'equipamento_apoi.e': 14,
            'equipamento_apoi.a': 15,
            'equipamento_apoi.peso': 16,
        })



def popular_acabamentos(sheet, acabamentos):
    # aceita estruturas diferentes: chave principal pode ser 'tabela_a'..'tabela_f'
    # Tabela A: linhas 8..27
    rows_a = list(range(8, 28))
    tabela_a = acabamentos.get('tabela_a', acabamentos.get('pisos', []))
    ra = 0
    for reg in tabela_a:
        if ra >= len(rows_a):
            break
        r = rows_a[ra]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Piso
        set_cell(sheet, r, 5, reg.get('piso', {}).get('tipo'))
        set_cell(sheet, r, 6, reg.get('piso', {}).get('dimensoes', {}).get('e'))
        set_cell(sheet, r, 7, reg.get('piso', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 8, reg.get('piso', {}).get('dimensoes', {}).get('l'))
        set_cell(sheet, r, 9, reg.get('piso', {}).get('dimensoes', {}).get('a'))
        set_cell(sheet, r, 10, reg.get('piso', {}).get('placa_ceramica', {}).get('c'))
        set_cell(sheet, r, 11, reg.get('piso', {}).get('placa_ceramica', {}).get('l'))
        # Soleiras
        set_cell(sheet, r, 12, reg.get('soleiras', {}).get('tipo'))
        set_cell(sheet, r, 13, reg.get('soleiras', {}).get('dimensoes', {}).get('e'))
        set_cell(sheet, r, 14, reg.get('soleiras', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 15, reg.get('soleiras', {}).get('dimensoes', {}).get('l'))
        # Rodapés
        set_cell(sheet, r, 16, reg.get('rodapes', {}).get('tipo'))
        set_cell(sheet, r, 17, reg.get('rodapes', {}).get('dimensoes', {}).get('h'))
        set_cell(sheet, r, 18, reg.get('rodapes', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 19, reg.get('rodapes', {}).get('dimensoes', {}).get('l'))
        set_cell(sheet, r, 20, reg.get('rodapes', {}).get('dimensoes', {}).get('a'))
        ra += 1

    # Tabela B: linhas 35..54
    rows_b = list(range(35, 55))
    tabela_b = acabamentos.get('tabela_b', [])
    rb = 0
    for reg in tabela_b:
        if rb >= len(rows_b):
            break
        r = rows_b[rb]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Azulejos e Rodabancas
        set_cell(sheet, r, 5, reg.get('azulejos', {}).get('tipo'))
        set_cell(sheet, r, 6, reg.get('azulejos', {}).get('dimensoes', {}).get('e'))
        set_cell(sheet, r, 7, reg.get('azulejos', {}).get('dimensoes', {}).get('h'))
        set_cell(sheet, r, 8, reg.get('azulejos', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 9, reg.get('azulejos', {}).get('dimensoes', {}).get('a'))
        set_cell(sheet, r, 10, reg.get('azulejos', {}).get('placa', {}).get('l'))
        set_cell(sheet, r, 11, reg.get('azulejos', {}).get('placa', {}).get('c'))
        # Peitoris
        set_cell(sheet, r, 12, reg.get('peitoris', {}).get('tipo'))
        set_cell(sheet, r, 13, reg.get('peitoris', {}).get('dimensoes', {}).get('e'))
        set_cell(sheet, r, 14, reg.get('peitoris', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 15, reg.get('peitoris', {}).get('dimensoes', {}).get('l'))
        # Forros
        set_cell(sheet, r, 16, reg.get('forros', {}).get('tipo'))
        set_cell(sheet, r, 17, reg.get('forros', {}).get('dimensoes', {}).get('l'))
        set_cell(sheet, r, 18, reg.get('forros', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 19, reg.get('forros', {}).get('dimensoes', {}).get('a'))
        set_cell(sheet, r, 20, reg.get('forros', {}).get('placa', {}).get('l'))
        set_cell(sheet, r, 21, reg.get('forros', {}).get('placa', {}).get('c'))
        rb += 1

    # Tabela C: linhas 61..80
    rows_c = list(range(61, 81))
    tabela_c = acabamentos.get('tabela_c', [])
    rc = 0
    for reg in tabela_c:
        if rc >= len(rows_c):
            break
        r = rows_c[rc]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Emassamento
        set_cell(sheet, r, 5, reg.get('emassamento', {}).get('h'))
        set_cell(sheet, r, 6, reg.get('emassamento', {}).get('per'))
        set_cell(sheet, r, 7, reg.get('emassamento', {}).get('a', {}).get('parede'))
        set_cell(sheet, r, 8, reg.get('emassamento', {}).get('a', {}).get('teto'))
        # Lixamento
        set_cell(sheet, r, 9, reg.get('lixamento', {}).get('h'))
        set_cell(sheet, r, 10, reg.get('lixamento', {}).get('per'))
        set_cell(sheet, r, 11, reg.get('lixamento', {}).get('a', {}).get('parede'))
        set_cell(sheet, r, 12, reg.get('lixamento', {}).get('a', {}).get('teto'))
        # Selamento
        set_cell(sheet, r, 13, reg.get('selamento', {}).get('h'))
        set_cell(sheet, r, 14, reg.get('selamento', {}).get('per'))
        set_cell(sheet, r, 15, reg.get('selamento', {}).get('a', {}).get('parede'))
        set_cell(sheet, r, 16, reg.get('selamento', {}).get('a', {}).get('teto'))
        # Pintura Acrílica
        set_cell(sheet, r, 17, reg.get('pintura_acrilica', {}).get('a', {}).get('parede'))
        set_cell(sheet, r, 18, reg.get('pintura_acrilica', {}).get('a', {}).get('teto'))
        set_cell(sheet, r, 19, reg.get('pintura_acrilica', {}).get('a', {}).get('piso'))
        set_cell(sheet, r, 20, reg.get('pintura_acrilica', {}).get('a', {}).get('pilar'))
        # Pintura Esmalte
        set_cell(sheet, r, 21, reg.get('pintura_esmalte', {}).get('a', {}).get('portas'))
        set_cell(sheet, r, 22, reg.get('pintura_esmalte', {}).get('a', {}).get('janelas'))
        set_cell(sheet, r, 23, reg.get('pintura_esmalte', {}).get('a', {}).get('grades'))
        rc += 1

    # Tabela D: linhas 87..106
    rows_d = list(range(87, 107))
    tabela_d = acabamentos.get('tabela_d', [])
    rd = 0
    for reg in tabela_d:
        if rd >= len(rows_d):
            break
        r = rows_d[rd]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Portas e Alçapões
        set_cell(sheet, r, 5, reg.get('portas', {}).get('pa'))
        set_cell(sheet, r, 6, reg.get('portas', {}).get('qtd'))
        set_cell(sheet, r, 7, reg.get('portas', {}).get('l'))
        set_cell(sheet, r, 8, reg.get('portas', {}).get('h'))
        set_cell(sheet, r, 9, reg.get('portas', {}).get('e'))
        set_cell(sheet, r, 10, reg.get('portas', {}).get('a'))
        # Janelas e Visores
        set_cell(sheet, r, 11, reg.get('janelas', {}).get('jv'))
        set_cell(sheet, r, 12, reg.get('janelas', {}).get('qtd'))
        set_cell(sheet, r, 13, reg.get('janelas', {}).get('l'))
        set_cell(sheet, r, 14, reg.get('janelas', {}).get('h'))
        set_cell(sheet, r, 15, reg.get('janelas', {}).get('e'))
        set_cell(sheet, r, 16, reg.get('janelas', {}).get('a'))
        # Telas
        set_cell(sheet, r, 17, reg.get('telas', {}).get('l'))
        set_cell(sheet, r, 18, reg.get('telas', {}).get('h'))
        set_cell(sheet, r, 19, reg.get('telas', {}).get('qtd'))
        set_cell(sheet, r, 20, reg.get('telas', {}).get('a'))
        # Vidros
        set_cell(sheet, r, 21, reg.get('vidros', {}).get('l'))
        set_cell(sheet, r, 22, reg.get('vidros', {}).get('h'))
        set_cell(sheet, r, 23, reg.get('vidros', {}).get('qtd'))
        set_cell(sheet, r, 24, reg.get('vidros', {}).get('a'))
        # Venezianas Industriais
        set_cell(sheet, r, 25, reg.get('venezianas', {}).get('peca'))
        set_cell(sheet, r, 26, reg.get('venezianas', {}).get('l'))
        set_cell(sheet, r, 27, reg.get('venezianas', {}).get('h'))
        set_cell(sheet, r, 28, reg.get('venezianas', {}).get('qtd'))
        set_cell(sheet, r, 29, reg.get('venezianas', {}).get('a'))
        # Protetores de Canto
        set_cell(sheet, r, 30, reg.get('protetores_canto', {}).get('l'))
        set_cell(sheet, r, 31, reg.get('protetores_canto', {}).get('e'))
        set_cell(sheet, r, 32, reg.get('protetores_canto', {}).get('c'))
        # Protetores de Parede
        set_cell(sheet, r, 33, reg.get('protetores_parede', {}).get('l'))
        set_cell(sheet, r, 34, reg.get('protetores_parede', {}).get('e'))
        set_cell(sheet, r, 35, reg.get('protetores_parede', {}).get('c'))
        # Grades
        set_cell(sheet, r, 36, reg.get('grades', {}).get('tipo'))
        set_cell(sheet, r, 37, reg.get('grades', {}).get('a'))
        set_cell(sheet, r, 38, reg.get('grades', {}).get('malha'))
        set_cell(sheet, r, 39, reg.get('grades', {}).get('e'))
        set_cell(sheet, r, 40, reg.get('grades', {}).get('afastamento', {}).get('janela'))
        set_cell(sheet, r, 41, reg.get('grades', {}).get('afastamento', {}).get('alvenaria'))
        rd += 1

    # Tabela E (Acessórios): linhas 113..132
    rows_e = list(range(113, 133))
    tabela_e = acabamentos.get('tabela_e', [])
    re = 0
    for reg in tabela_e:
        if re >= len(rows_e):
            break
        r = rows_e[re]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        set_cell(sheet, r, 5, reg.get('acessorios', {}).get('bacias'))
        set_cell(sheet, r, 6, reg.get('acessorios', {}).get('mictorios'))
        set_cell(sheet, r, 7, reg.get('acessorios', {}).get('lavatórios'))
        set_cell(sheet, r, 8, reg.get('acessorios', {}).get('cubas'))
        set_cell(sheet, r, 9, reg.get('acessorios', {}).get('tanques'))
        set_cell(sheet, r, 10, reg.get('acessorios', {}).get('torneiras'))
        # Barras de Apoio
        set_cell(sheet, r, 11, reg.get('barras_apoio', {}).get('qtd'))
        set_cell(sheet, r, 12, reg.get('barras_apoio', {}).get('c'))
        set_cell(sheet, r, 13, reg.get('barras_apoio', {}).get('diametro'))
        set_cell(sheet, r, 14, reg.get('barras_apoio', {}).get('h_piso'))
        # Corrimãos e Guarda-Corpos
        set_cell(sheet, r, 15, reg.get('corrimaos', {}).get('qtd'))
        set_cell(sheet, r, 16, reg.get('corrimaos', {}).get('c'))
        set_cell(sheet, r, 17, reg.get('corrimaos', {}).get('diametro'))
        set_cell(sheet, r, 18, reg.get('corrimaos', {}).get('h_piso'))
        re += 1

    # Tabela E (Bancadas e Pias / Divisórias / Boxes): linhas 140..159
    rows_e2 = list(range(140, 160))
    tabela_e2 = acabamentos.get('tabela_e2', acabamentos.get('bancadas', []))
    re2 = 0
    for reg in tabela_e2:
        if re2 >= len(rows_e2):
            break
        r = rows_e2[re2]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        # Bancadas e Pias
        set_cell(sheet, r, 5, reg.get('bancadas', {}).get('dimensoes', {}).get('a'))
        set_cell(sheet, r, 6, reg.get('bancadas', {}).get('dimensoes', {}).get('h'))
        set_cell(sheet, r, 7, reg.get('bancadas', {}).get('tampo', {}).get('l'))
        set_cell(sheet, r, 8, reg.get('bancadas', {}).get('tampo', {}).get('c'))
        set_cell(sheet, r, 9, reg.get('bancadas', {}).get('tampo', {}).get('e'))
        set_cell(sheet, r, 10, reg.get('bancadas', {}).get('frontao', {}).get('l'))
        set_cell(sheet, r, 11, reg.get('bancadas', {}).get('frontao', {}).get('c'))
        set_cell(sheet, r, 12, reg.get('bancadas', {}).get('frontao', {}).get('e'))
        set_cell(sheet, r, 13, reg.get('bancadas', {}).get('saia', {}).get('l'))
        set_cell(sheet, r, 14, reg.get('bancadas', {}).get('saia', {}).get('c'))
        set_cell(sheet, r, 15, reg.get('bancadas', {}).get('saia', {}).get('e'))
        # Divisórias
        set_cell(sheet, r, 16, reg.get('divisoes', {}).get('tipo'))
        set_cell(sheet, r, 17, reg.get('divisoes', {}).get('qtd'))
        set_cell(sheet, r, 18, reg.get('divisoes', {}).get('peca', {}).get('c'))
        set_cell(sheet, r, 19, reg.get('divisoes', {}).get('peca', {}).get('h'))
        set_cell(sheet, r, 20, reg.get('divisoes', {}).get('peca', {}).get('a'))
        set_cell(sheet, r, 21, reg.get('divisoes', {}).get('portas', {}).get('l'))
        set_cell(sheet, r, 22, reg.get('divisoes', {}).get('portas', {}).get('h'))
        set_cell(sheet, r, 23, reg.get('divisoes', {}).get('portas', {}).get('a'))
        # Boxes
        set_cell(sheet, r, 24, reg.get('boxes', {}).get('tipo'))
        set_cell(sheet, r, 25, reg.get('boxes', {}).get('qtd'))
        set_cell(sheet, r, 26, reg.get('boxes', {}).get('peca', {}).get('l'))
        set_cell(sheet, r, 27, reg.get('boxes', {}).get('peca', {}).get('c'))
        set_cell(sheet, r, 28, reg.get('boxes', {}).get('peca', {}).get('e'))
        set_cell(sheet, r, 29, reg.get('boxes', {}).get('sobreposicao'))
        set_cell(sheet, r, 30, reg.get('boxes', {}).get('a'))
        re2 += 1

    # Tabela F: Mobiliário - linhas 166..185
    rows_f = list(range(166, 186))
    tabela_f = acabamentos.get('tabela_f', [])
    rf = 0
    for reg in tabela_f:
        if rf >= len(rows_f):
            break
        r = rows_f[rf]
        set_cell(sheet, r, 2, reg.get('ambiente'))
        set_cell(sheet, r, 5, reg.get('mobiliario', {}).get('tipo'))
        set_cell(sheet, r, 6, reg.get('mobiliario', {}).get('dimensoes', {}).get('h'))
        set_cell(sheet, r, 7, reg.get('mobiliario', {}).get('dimensoes', {}).get('l'))
        set_cell(sheet, r, 8, reg.get('mobiliario', {}).get('dimensoes', {}).get('c'))
        set_cell(sheet, r, 9, reg.get('mobiliario', {}).get('a'))
        rf += 1


    def popular_alvenarias(sheet, alvenarias):
        """Preenche a aba 'Alvenarias' conforme mapeamento fornecido."""
        def fill_rows(start, end, items, mapping):
            rows = list(range(start, end + 1))
            for idx, it in enumerate(items):
                if idx >= len(rows):
                    break
                r = rows[idx]
                for key, col in mapping.items():
                    # suporta chaves aninhadas com ponto
                    cur = it
                    val = None
                    for part in key.split('.'):
                        if isinstance(cur, dict):
                            cur = cur.get(part)
                        else:
                            cur = None
                            break
                    val = cur
                    set_cell(sheet, r, col, val)

        # Parte 1: rows 10..29
        part1 = alvenarias.get('parte1', [])
        fill_rows(10, 29, part1, {
            'ambiente': 2,
            # Painéis em Alvenaria
            'painel_alvenaria.peca': 5, 'painel_alvenaria.c': 6, 'painel_alvenaria.l': 7, 'painel_alvenaria.h': 8,
            'painel_alvenaria.vaos': 9, 'painel_alvenaria.a': 10,
            # Gesso Acartonado
            'gesso.peca': 11, 'gesso.c': 12, 'gesso.l': 13, 'gesso.h': 14, 'gesso.vaos': 15, 'gesso.a': 16,
            # Cobogó / Blocos de vidro
            'cobogo.peca': 17, 'cobogo.c': 18, 'cobogo.l': 19, 'cobogo.h': 20, 'cobogo.vaos': 21,
            'cobogo.a1': 22, 'cobogo.a2': 23, 'cobogo.a3': 24, 'cobogo.a4': 25
        })

        # Parte 2: rows 36..55
        part2 = alvenarias.get('parte2', [])
        fill_rows(36, 55, part2, {
            'ambiente': 2, 'peca': 5, 'qtd.verga': 6, 'qtd.c_verga': 7,
            'l': 8, 'c': 9, 'h': 10, 'engastamento': 11, 'concreto': 12, 'ferragem': 13
        })

        # Parte 3: rows 63..81
        part3 = alvenarias.get('parte3', [])
        fill_rows(63, 81, part3, {
            'ambiente': 2,
            'guias.local': 5, 'guias.l': 6, 'guias.h': 7, 'guias.c': 8, 'guias.concreto': 9,
            'calcadas.local': 10, 'calcadas.l': 11, 'calcadas.h': 12, 'calcadas.e': 13, 'calcadas.c': 14, 'calcadas.a': 15, 'calcadas.concreto': 16
        })

        # Parte 4: rows 89..108
        part4 = alvenarias.get('parte4', [])
        fill_rows(89, 108, part4, {
            'ambiente': 2, 'local': 5,
            'dimensoes.c_proj': 6, 'dimensoes.c_real': 7, 'dimensoes.l': 8, 'dimensoes.h': 9, 'dimensoes.i_pct': 10,
            'piso.e': 11, 'piso.a': 12, 'piso.concreto': 13,
            'parede_contencao.l': 14, 'parede_contencao.h': 15, 'parede_contencao.c': 16, 'parede_contencao.a': 17, 'parede_contencao.concreto': 18,
            'guia_balizamento.l': 19, 'guia_balizamento.h': 20,
            'armacao.ferragem': 21, 'armacao.estribo': 22,
            'forma_madeira.h': 23, 'forma_madeira.c': 24, 'forma_madeira.a': 25
        })

        # Parte 5: rows 140..157
        part5 = alvenarias.get('parte5', [])
        fill_rows(140, 157, part5, {
            'ambiente': 2, 'local': 5, 'peca': 6,
            'material_fechamento.c': 7, 'material_fechamento.h': 8, 'material_fechamento.a': 9,
            'mouroes.qtd': 10, 'mouroes.c': 11, 'mouroes.h': 12,
            'esticador.qtd': 13, 'esticador.c': 14, 'esticador.h': 15,
            'concertina.c': 16
        })

        # Parte 6: rows 165..184
        part6 = alvenarias.get('parte6', [])
        fill_rows(165, 184, part6, {
            'ambiente': 2, 'local': 5,
            'apicoamento.h': 6, 'apicoamento.c': 7, 'apicoamento.a': 8,
            'chapisco.h': 9, 'chapisco.c': 10, 'chapisco.a': 11,
            'emboço_pintura.h': 12, 'emboço_pintura.c': 13, 'emboço_pintura.a': 14,
            'emboço_revestimento.h': 15, 'emboço_revestimento.c': 16, 'emboço_revestimento.a': 17
        })

        # Parte 7: rows 192..211
        part7 = alvenarias.get('parte7', [])
        fill_rows(192, 211, part7, {
            'ambiente': 2, 'local': 5,
            'lastros.l': 6, 'lastros.c': 7, 'lastros.e': 8, 'lastros.v': 9,
            'contrapisos.l': 10, 'contrapisos.e': 11, 'contrapisos.c': 12, 'contrapisos.v': 13, 'contrapisos.ferragem': 14,
            'juntas_dilatacao.l': 15, 'juntas_dilatacao.h': 16, 'juntas_dilatacao.c': 17
        })

        # Parte 9: rows 245..264
        part9 = alvenarias.get('parte9', [])
        fill_rows(245, 264, part9, {
            'ambiente': 2, 'local': 5,
            'estruturas.l': 6, 'estruturas.h': 7,
            'estruturas.secao.c': 8, 'estruturas.secao.l': 9, 'estruturas.secao.h': 10, 'estruturas.a_tot': 11,
            'pisos.secao.c': 12, 'pisos.secao.l': 13, 'pisos.secao.h': 14, 'pisos.per': 15, 'pisos.a_tot': 16,
            'paredes.secao.c': 17, 'paredes.secao.l': 18, 'paredes.secao.h': 19, 'paredes.per': 20, 'paredes.a_tot': 21
        })



def popular_instalacoes(sheet, instalacoes):
    # Instalações elétricas
    sheet.cell(row=1, column=2, value='Instalações Elétricas - Local')
    r = 2
    for it in instalacoes.get('inst_eletricas', []):
        sheet.cell(row=r, column=2, value=it.get('local'))
        sheet.cell(row=r, column=3, value=it.get('circuito'))
        sheet.cell(row=r, column=4, value=it.get('qtd_cabos'))
        sheet.cell(row=r, column=5, value=it.get('cabo_especificacao'))
        r += 1


def popular_comunicacoes(sheet, lista):
    sheet.cell(row=1, column=2, value='Ambiente')
    sheet.cell(row=1, column=3, value='Local')
    sheet.cell(row=1, column=4, value='Extintor')
    r = 2
    for it in lista:
        sheet.cell(row=r, column=2, value=it.get('ambiente'))
        sheet.cell(row=r, column=3, value=it.get('local'))
        sheet.cell(row=r, column=4, value=it.get('extintor'))
        r += 1


def popular_excel(caminho_excel, dados):
    caminho_excel = Path(caminho_excel)
    if caminho_excel.exists():
        workbook = openpyxl.load_workbook(caminho_excel)
    else:
        workbook = openpyxl.Workbook()
        # remove sheet default vazio
        if 'Sheet' in workbook.sheetnames:
            std = workbook['Sheet']
            workbook.remove(std)

    # Levantamento Campo
    sheet_l = _criar_sheet(workbook, 'Levantamento Campo')
    levantamento_campo = dados.get('levantamento_campo', dados)
    if not isinstance(levantamento_campo, dict):
        levantamento_campo = dados
    popular_levantamento_campo(sheet_l, levantamento_campo)

    # Serviços preliminares
    sheet_sp = _criar_sheet(workbook, 'Serviços Preliminares')
    popular_servicos_preliminares(sheet_sp, dados.get('servicos_preliminares', []))

    # Movimento de solo
    sheet_ms = _criar_sheet(workbook, 'Movimento de Solo')
    popular_movimento_solo(sheet_ms, dados.get('movimento_solo', []))

    # Estruturas
    sheet_e = _criar_sheet(workbook, 'Estruturas')
    popular_estruturas(sheet_e, dados.get('estruturas', {}))

    # Acabamentos
    sheet_a = _criar_sheet(workbook, 'Acabamentos')
    popular_acabamentos(sheet_a, dados.get('acabamentos', {}))

    # Alvenarias
    sheet_al = _criar_sheet(workbook, 'Alvenarias')
    popular_alvenarias(sheet_al, dados.get('alvenarias', {}))

    # Instalações
    sheet_i = _criar_sheet(workbook, 'Instalacoes')
    popular_instalacoes(sheet_i, dados.get('instalacoes', {}))

    # Comunicações Ambientais
    sheet_c = _criar_sheet(workbook, 'Comunicacoes')
    popular_comunicacoes(sheet_c, dados.get('comunicacoes_ambientais', []))

    # Salva
    workbook.save(caminho_excel)
    print(f"Planilha atualizada: {caminho_excel}")


def main():
    base = Path(__file__).parent
    caminho_json = base / 'dados_levantamento_teste.json'

    # procurar template na pasta atual, na pasta pai e na raiz do workspace
    # suportar variações de nome como 'memorial_model.xlsx' e procura recursiva por arquivos contendo 'memorial'
    candidates = [
        base / 'model_memorial.xlsx',
        base / 'memorial_model.xlsx',
        base.parent / 'model_memorial.xlsx',
        base.parent / 'memorial_model.xlsx',
        base.parent.parent / 'model_memorial.xlsx',
        base.parent.parent / 'memorial_model.xlsx',
    ]

    # acrescenta resultados de busca recursiva (rglob) nas pastas próximas
    for root in (base, base.parent, base.parent.parent):
        try:
            for p in root.rglob('*memorial*.xlsx'):
                candidates.append(p)
        except Exception:
            # em casos raros root pode ser None ou inacessível
            pass

    caminho_template = next((p for p in candidates if p.exists()), None)

    if caminho_template is None:
        print("Arquivo de template 'model_memorial.xlsx' não encontrado nas pastas esperadas.")
        print(f"Verifique onde o arquivo está e mova para: {base} ou {base.parent}")
        return

    dados = carregar_dados_json(caminho_json)

    # criar backup antes de sobrescrever
    backup_path = caminho_template.with_name(caminho_template.stem + '_backup' + caminho_template.suffix)
    try:
        shutil.copy2(str(caminho_template), str(backup_path))
        print(f"Backup criado: {backup_path}")
    except Exception as e:
        print(f"Falha ao criar backup: {e}")

    # carrega o workbook do template e popula todas as abas relevantes
    workbook = openpyxl.load_workbook(str(caminho_template))

    # Levantamento Campo
    sheet_l = _criar_sheet(workbook, 'Levantamento Campo')
    popular_levantamento_campo(sheet_l, dados)

    # Serviços Preliminares
    sheet_sp = _criar_sheet(workbook, 'Serviços Preliminares')
    popular_servicos_preliminares(sheet_sp, dados.get('servicos_preliminares', {}))

    # Movimento de Solo
    sheet_ms = _criar_sheet(workbook, 'Movimento de Solo')
    # permite tanto dict com subtabelas quanto listas
    popular_movimento_solo(sheet_ms, dados.get('movimento_solo', {}))

    # Estruturas
    sheet_e = _criar_sheet(workbook, 'Estruturas')
    popular_estruturas(sheet_e, dados.get('estruturas', {}))

    # Acabamentos
    sheet_a = _criar_sheet(workbook, 'Acabamentos')
    popular_acabamentos(sheet_a, dados.get('acabamentos', {}))

    # Alvenarias
    sheet_al = _criar_sheet(workbook, 'Alvenarias')
    popular_alvenarias(sheet_al, dados.get('alvenarias', {}))

    # Instalações
    sheet_i = _criar_sheet(workbook, 'Instalacoes')
    popular_instalacoes(sheet_i, dados.get('instalacoes', {}))

    # Comunicações Ambientais
    sheet_c = _criar_sheet(workbook, 'Comunicacoes')
    popular_comunicacoes(sheet_c, dados.get('comunicacoes_ambientais', []))

    # salva diretamente no arquivo template (sobrescreve)
    try:
        workbook.save(str(caminho_template))
        print(f"Template atualizado: {caminho_template}")
    except PermissionError:
        # provavelmente o arquivo está aberto no Excel — salvar cópia alternativa
        alt_path = caminho_template.with_name(caminho_template.stem + '_out' + caminho_template.suffix)
        try:
            workbook.save(str(alt_path))
            print(f"Arquivo estava em uso. Salvo como cópia: {alt_path}")
        except Exception as e:
            print(f"Falha ao salvar arquivo alternativo: {e}")
    except Exception as e:
        print(f"Erro ao salvar template: {e}")


if __name__ == "__main__":
    main()
