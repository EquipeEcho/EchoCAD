# generatorteste.py
# Gera o memorial de cálculo a partir dos dados extraídos do DXF.
# Cria um template novo e bem estruturado (não depende do model_memorial.xlsx
# para as colunas de dados), mas mantém a aba "Levantamento Campo" compatível.

import logging
import math
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .dxf_extractor import Ambiente, CADExtractor, ProjetoMemorial
from .sinapi import buscar_preco_sinapi, carregar_sinapi

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Mapeamento ambiente → material SINAPI
# ---------------------------------------------------------------------------
MAP_SINAPI = {
    "alojamento": "bloco cerâmico",
    "banheiro": "bloco cerâmico",
    "copa": "bloco cerâmico",
    "sala": "bloco cerâmico",
    "circulação": "bloco cerâmico",
    "auditório": "bloco cerâmico",
    "reserva": "bloco cerâmico",
    "passadiço": "bloco cerâmico",
    "calçada": "concreto estrutural",
    "área": "concreto estrutural",
    "telhado": "telha cerâmica",
}


def _categoria_sinapi(nome: str) -> str:
    nome_l = nome.lower()
    for chave, material in MAP_SINAPI.items():
        if chave in nome_l:
            return material
    return "bloco cerâmico"


def _criar_ambiente_por_quantitativos(quantitativos: dict[str, Any] | None):
    if not quantitativos:
        return []

    paredes = quantitativos.get("paredes", [])
    if not paredes:
        return []

    resumo = quantitativos.get("resumo_global", {})
    total_comprimento = sum(p.get("comprimento_m") or 0 for p in paredes)
    total_area_bruta = sum(p.get("area_externa_m2") or 0 for p in paredes)
    total_volume_bruto = sum(p.get("volume_bruto_m3") or 0 for p in paredes)
    total_volume_liquido = resumo.get("volume_final_liquido_alvenaria_m3")
    total_descontos = sum(
        d.get("volume_descontado_m3") or 0
        for p in paredes
        for d in p.get("descontos_aberturas", [])
    )
    espessuras = [p.get("espessura_m") for p in paredes if p.get("espessura_m")]
    espessura = sum(espessuras) / len(espessuras) if espessuras else 0.15

    if total_comprimento <= 0 or total_area_bruta <= 0:
        return []

    pe_direito = total_area_bruta / total_comprimento
    area_vaos = total_descontos / espessura if espessura else 0
    if not area_vaos and total_volume_bruto and total_volume_liquido is not None:
        area_vaos = max((total_volume_bruto - total_volume_liquido) / espessura, 0)

    area_total = resumo.get("area_total_laje_m2") or 0
    ambiente = Ambiente(
        nome="AMBIENTE TECNICO",
        subtitulo="Gerado pelos quantitativos do DXF",
        area=round(area_total, 2),
        perimetro=round(total_comprimento, 2),
        pe_direito=round(pe_direito, 2),
        espessura_parede=round(espessura, 3),
        comprimento_paredes=round(total_comprimento, 2),
        comprimento_vaos=round(area_vaos / CADExtractor.ALTURA_VAO_PADRAO, 2),
        area_bruta_parede=round(total_area_bruta, 2),
        area_vaos=round(area_vaos, 2),
        area_liquida_parede=round(max(total_area_bruta - area_vaos, 0), 2),
    )
    return [ambiente]


def _media_valores_parede(paredes: list[dict[str, Any]], chave: str, padrao: float) -> float:
    valores = [p.get(chave) for p in paredes if p.get(chave)]
    return sum(valores) / len(valores) if valores else padrao


def _normalizar_ponto(ponto, casas: int = 6):
    return (round(float(ponto[0]), casas), round(float(ponto[1]), casas))


def _distancia(p1, p2) -> float:
    return math.hypot(p2[0] - p1[0], p2[1] - p1[1])


def _area_poligono(vertices: list[tuple[float, float]]) -> float:
    if len(vertices) < 3:
        return 0.0
    return abs(
        sum(
            vertices[i][0] * vertices[(i + 1) % len(vertices)][1]
            - vertices[(i + 1) % len(vertices)][0] * vertices[i][1]
            for i in range(len(vertices))
        )
        / 2
    )


def _ordenar_componente_fechado(segmentos: list[tuple[tuple[float, float], tuple[float, float]]]):
    if len(segmentos) < 3:
        return []

    adj: dict[tuple[float, float], list[tuple[float, float]]] = {}
    for p1, p2 in segmentos:
        adj.setdefault(p1, []).append(p2)
        adj.setdefault(p2, []).append(p1)

    if any(len(vizinhos) != 2 for vizinhos in adj.values()):
        return []

    inicio = segmentos[0][0]
    anterior = None
    atual = inicio
    vertices = [inicio]

    for _ in range(len(segmentos)):
        candidatos = [p for p in adj[atual] if p != anterior]
        if not candidatos:
            return []
        proximo = candidatos[0]
        anterior, atual = atual, proximo
        if atual == inicio:
            return vertices
        vertices.append(atual)

    return []


def _extrair_contornos_fechados_dxf(dxf_file: str):
    try:
        import ezdxf
    except Exception as exc:
        logger.warning(f"Nao foi possivel importar ezdxf para fallback geometrico: {exc}")
        return []

    try:
        doc = ezdxf.readfile(dxf_file)
    except Exception as exc:
        logger.warning(f"Nao foi possivel ler DXF para fallback geometrico: {exc}")
        return []

    msp = doc.modelspace()
    segmentos: list[tuple[tuple[float, float], tuple[float, float]]] = []
    contornos: list[list[tuple[float, float]]] = []

    for entidade in msp:
        layer = entidade.dxf.layer.upper()
        if "PAREDE" not in layer and "ALVENARIA" not in layer:
            continue

        if entidade.dxftype() == "LINE":
            p1 = _normalizar_ponto((entidade.dxf.start.x, entidade.dxf.start.y))
            p2 = _normalizar_ponto((entidade.dxf.end.x, entidade.dxf.end.y))
            if p1 != p2:
                segmentos.append((p1, p2))
        elif entidade.dxftype() == "LWPOLYLINE":
            pontos = [_normalizar_ponto((x, y)) for x, y, *_ in entidade.get_points()]
            if entidade.closed and len(pontos) >= 3:
                contornos.append(pontos)

    restantes = segmentos[:]
    while restantes:
        pilha = [restantes.pop(0)]
        componente = []
        pontos_componente = set(pilha[0])

        while pilha:
            seg = pilha.pop()
            componente.append(seg)

            mudou = True
            while mudou:
                mudou = False
                for idx, candidato in list(enumerate(restantes)):
                    if candidato[0] in pontos_componente or candidato[1] in pontos_componente:
                        restantes.pop(idx)
                        pilha.append(candidato)
                        pontos_componente.update(candidato)
                        mudou = True
                        break

        vertices = _ordenar_componente_fechado(componente)
        if vertices:
            contornos.append(vertices)

    contornos_validos = []
    for vertices in contornos:
        area = _area_poligono(vertices)
        perimetro = sum(
            _distancia(vertices[i], vertices[(i + 1) % len(vertices)])
            for i in range(len(vertices))
        )
        if area > 0 and perimetro > 0:
            contornos_validos.append((area, perimetro, vertices))

    contornos_validos.sort(key=lambda item: item[0], reverse=True)
    return contornos_validos


def _criar_ambientes_por_contornos(dxf_file: str, quantitativos: dict[str, Any] | None):
    contornos = _extrair_contornos_fechados_dxf(dxf_file)
    if len(contornos) < 2:
        return []

    paredes = (quantitativos or {}).get("paredes", [])
    pe_direito = _media_valores_parede(paredes, "altura_m", 2.8)
    espessura = _media_valores_parede(paredes, "espessura_m", 0.15)

    ambientes = []
    for idx, (area, perimetro, vertices) in enumerate(contornos, start=1):
        area_bruta = perimetro * pe_direito
        cx = sum(p[0] for p in vertices) / len(vertices)
        cy = sum(p[1] for p in vertices) / len(vertices)
        ambientes.append(
            Ambiente(
                nome=f"AMBIENTE {idx}",
                subtitulo="Gerado por contorno fechado do DXF",
                area=round(area, 2),
                perimetro=round(perimetro, 2),
                pe_direito=round(pe_direito, 2),
                espessura_parede=round(espessura, 3),
                comprimento_paredes=round(perimetro, 2),
                area_bruta_parede=round(area_bruta, 2),
                area_vaos=0,
                area_liquida_parede=round(area_bruta, 2),
                cx=cx,
                cy=cy,
            )
        )

    return ambientes


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
COR_HEADER_AZUL = "FF003366"  # azul escuro Exército
COR_HEADER_CINZA = "FFD9D9D9"
COR_LINHA_PAR = "FFF5F5F5"
COR_TOTAL = "FFBDD7EE"
FONTE_TITULO = Font(name="Arial", bold=True, size=14, color="FFFFFFFF")
FONTE_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFFFF")
FONTE_SUBHEADER = Font(name="Arial", bold=True, size=9, color="FF000000")
FONTE_DADO = Font(name="Arial", size=9)
FONTE_DADO_BOLD = Font(name="Arial", bold=True, size=9)
FONTE_TOTAL = Font(name="Arial", bold=True, size=9, color="FF003366")

FILL_AZUL = PatternFill("solid", fgColor=COR_HEADER_AZUL)
FILL_CINZA = PatternFill("solid", fgColor=COR_HEADER_CINZA)
FILL_PAR = PatternFill("solid", fgColor=COR_LINHA_PAR)
FILL_TOTAL = PatternFill("solid", fgColor=COR_TOTAL)

BORDA_FINA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BORDA_MEDIA = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

ALINHAR_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHAR_ESQ = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALINHAR_DIR = Alignment(horizontal="right", vertical="center")

FMT_NUMERO = "#,##0.00"
FMT_MOEDA = "R$ #,##0.00"
FMT_INTEIRO = "#,##0"


def _aplicar_borda(cell, borda=BORDA_FINA):
    cell.border = borda


def _celula(
    ws,
    row,
    col,
    valor=None,
    fonte=None,
    fill=None,
    alinhamento=None,
    formato=None,
    borda=BORDA_FINA,
):
    c = ws.cell(row=row, column=col, value=valor)
    if fonte:
        c.font = fonte
    if fill:
        c.fill = fill
    if alinhamento:
        c.alignment = alinhamento
    if formato:
        c.number_format = formato
    if borda:
        c.border = borda
    return c


# ---------------------------------------------------------------------------
# Gerador do memorial
# ---------------------------------------------------------------------------
class MemorialGenerator:
    # Colunas da aba "Levantamento Campo"
    # B  = Ambiente
    # C  = Subtítulo/uso
    # D  = Área (m²)
    # E  = Perímetro (m)
    # F  = Pé-direito (m)
    # G  = Área bruta de parede (m²)
    # H  = Área de vãos (m²)
    # I  = Área líquida de parede (m²)
    # J  = Espessura parede (m)
    # K  = Material / categoria
    # L  = Preço unitário SINAPI (R$/m²)
    # M  = Custo total (R$)

    COL_NOME = 2  # B
    COL_SUBTITULO = 3  # C
    COL_AREA = 4  # D
    COL_PERIMETRO = 5  # E
    COL_PD = 6  # F
    COL_AREA_BRUTA = 7  # G
    COL_AREA_VAO = 8  # H
    COL_AREA_LIQ = 9  # I
    COL_ESP = 10  # J
    COL_MATERIAL = 11  # K
    COL_PU = 12  # L
    COL_TOTAL = 13  # M

    ROW_TITULO = 1
    ROW_SUBTITULO = 2
    ROW_HEADER1 = 4
    ROW_HEADER2 = 5
    ROW_DADOS_INI = 6

    def __init__(self, template_path: str = None):
        # template_path é aceito mas ignorado — geramos nosso próprio template
        pass

    def _criar_workbook(
        self, projeto: ProjetoMemorial, quantitativos: dict[str, Any] | None = None
    ) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Levantamento Campo"

        self._preencher_aba_levantamento(ws, projeto)
        self._ajustar_larguras(ws)
        if quantitativos and "erro" not in quantitativos:
            self._preencher_aba_resumo_dxf(wb, quantitativos)
            self._preencher_aba_paredes(wb, quantitativos)
            self._preencher_aba_estruturas(wb, quantitativos)
            self._preencher_aba_infraestrutura(wb, quantitativos)

        return wb

    @staticmethod
    def _nova_aba(wb: Workbook, nome: str):
        if nome in wb.sheetnames:
            del wb[nome]
        return wb.create_sheet(title=nome)

    @staticmethod
    def _escrever_titulo(ws, row: int, col_ini: int, col_fim: int, titulo: str):
        ws.merge_cells(start_row=row, start_column=col_ini, end_row=row, end_column=col_fim)
        cell = ws.cell(row=row, column=col_ini, value=titulo)
        cell.font = FONTE_TITULO
        cell.fill = FILL_AZUL
        cell.alignment = ALINHAR_CENTRO
        cell.border = BORDA_MEDIA

    @staticmethod
    def _escrever_tabela(ws, row: int, headers: list[str], rows: list[list[Any]]) -> int:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = FONTE_HEADER
            cell.fill = FILL_AZUL
            cell.alignment = ALINHAR_CENTRO
            cell.border = BORDA_FINA

        for ridx, values in enumerate(rows, start=row + 1):
            fill = FILL_PAR if (ridx - row) % 2 == 1 else None
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=ridx, column=col, value=value)
                cell.font = FONTE_DADO
                cell.alignment = ALINHAR_CENTRO
                cell.border = BORDA_FINA
                if fill:
                    cell.fill = fill
                if isinstance(value, float):
                    cell.number_format = FMT_NUMERO

        return row + len(rows) + 2

    @staticmethod
    def _ajustar_larguras_por_headers(ws, headers: list[str]):
        for idx, header in enumerate(headers, start=1):
            largura = max(12, min(max(len(header) + 4, 16), 34))
            ws.column_dimensions[get_column_letter(idx)].width = largura

    def _preencher_aba_resumo_dxf(self, wb: Workbook, quantitativos: dict[str, Any]):
        ws = self._nova_aba(wb, "Resumo DXF")
        resumo = quantitativos.get("resumo_global", {})
        self._escrever_titulo(ws, 1, 1, 3, "RESUMO DOS QUANTITATIVOS DO DXF")

        labels = {
            "quantidade_total_portas": "Portas",
            "quantidade_total_janelas": "Janelas",
            "volume_total_descontado_vãos_m3": "Volume descontado de vaos (m3)",
            "volume_final_liquido_alvenaria_m3": "Volume liquido de alvenaria (m3)",
            "quantidade_total_colunas": "Colunas",
            "volume_total_vigas_m3": "Volume de vigas (m3)",
            "volume_total_colunas_m3": "Volume de colunas (m3)",
            "area_total_laje_m2": "Area de laje (m2)",
            "volume_total_laje_m3": "Volume de laje (m3)",
            "comprimento_total_fios_m": "Comprimento de fios (m)",
            "comprimento_total_canos_m": "Comprimento de canos (m)",
        }
        rows = [[label, resumo.get(key, 0)] for key, label in labels.items()]
        headers = ["Item", "Valor"]
        self._escrever_tabela(ws, 3, headers, rows)
        self._ajustar_larguras_por_headers(ws, headers)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18

    def _preencher_aba_paredes(self, wb: Workbook, quantitativos: dict[str, Any]):
        paredes = quantitativos.get("paredes", [])
        if not paredes:
            return

        ws = self._nova_aba(wb, "Paredes")
        self._escrever_titulo(ws, 1, 1, 9, "PAREDES E DESCONTOS DE ABERTURAS")
        headers = [
            "ID",
            "Layer",
            "Comprimento (m)",
            "Altura (m)",
            "Espessura (m)",
            "Area externa (m2)",
            "Volume bruto (m3)",
            "Descontos",
            "Volume liquido (m3)",
        ]
        rows = []
        for parede in paredes:
            descontos = parede.get("descontos_aberturas", [])
            descontos_txt = "; ".join(
                f"{d.get('tipo')} {d.get('especificacao')} (-{d.get('volume_descontado_m3')} m3)"
                for d in descontos
            )
            rows.append(
                [
                    parede.get("id_dxf"),
                    parede.get("layer"),
                    parede.get("comprimento_m"),
                    parede.get("altura_m"),
                    parede.get("espessura_m"),
                    parede.get("area_externa_m2"),
                    parede.get("volume_bruto_m3"),
                    descontos_txt,
                    parede.get("volume_liquido_m3"),
                ]
            )

        self._escrever_tabela(ws, 3, headers, rows)
        self._ajustar_larguras_por_headers(ws, headers)
        ws.column_dimensions["H"].width = 36

    def _preencher_aba_estruturas(self, wb: Workbook, quantitativos: dict[str, Any]):
        ws = self._nova_aba(wb, "Estruturas DXF")
        self._escrever_titulo(ws, 1, 1, 8, "ESTRUTURAS EXTRAIDAS DO DXF")

        row = 3
        vigas = quantitativos.get("vigas", [])
        if vigas:
            row = self._escrever_tabela(
                ws,
                row,
                ["Tipo", "Origem", "Layer", "Comprimento (m)", "Base (m)", "Altura (m)", "Volume (m3)"],
                [
                    [
                        v.get("tipo") or v.get("layer"),
                        v.get("origem", "LINE"),
                        v.get("layer"),
                        v.get("comprimento_m"),
                        v.get("base_m"),
                        v.get("altura_m"),
                        v.get("volume_m3"),
                    ]
                    for v in vigas
                ],
            )

        colunas = quantitativos.get("colunas", [])
        if colunas:
            row = self._escrever_tabela(
                ws,
                row,
                ["Tipo", "X", "Y", "Largura (m)", "Profundidade (m)", "Altura (m)", "Volume (m3)"],
                [
                    [
                        c.get("tipo"),
                        (c.get("coordenada") or [None, None])[0],
                        (c.get("coordenada") or [None, None])[1],
                        c.get("largura_m"),
                        c.get("profundidade_m"),
                        c.get("altura_m"),
                        c.get("volume_m3"),
                    ]
                    for c in colunas
                ],
            )

        lajes = quantitativos.get("laje", [])
        if lajes:
            self._escrever_tabela(
                ws,
                row,
                ["Layer", "Origem", "Area (m2)", "Espessura (m)", "Volume (m3)"],
                [
                    [
                        l.get("layer"),
                        l.get("origem"),
                        l.get("area_m2"),
                        l.get("espessura_m"),
                        l.get("volume_m3"),
                    ]
                    for l in lajes
                ],
            )

        self._ajustar_larguras_por_headers(
            ws,
            ["Tipo", "Origem", "Layer", "Comprimento (m)", "Base (m)", "Altura (m)", "Volume (m3)"],
        )

    def _preencher_aba_infraestrutura(self, wb: Workbook, quantitativos: dict[str, Any]):
        infra = quantitativos.get("infraestrutura", {})
        eletrica = infra.get("eletrica", [])
        hidraulica = infra.get("hidraulica", [])
        if not eletrica and not hidraulica:
            return

        ws = self._nova_aba(wb, "Infraestrutura")
        self._escrever_titulo(ws, 1, 1, 3, "INFRAESTRUTURA EXTRAIDA DO DXF")
        rows = [
            ["Eletrica", item.get("tipo"), item.get("comprimento_m")]
            for item in eletrica
        ]
        rows.extend(
            ["Hidraulica", item.get("tipo"), item.get("comprimento_m")]
            for item in hidraulica
        )
        headers = ["Sistema", "Tipo", "Comprimento (m)"]
        self._escrever_tabela(ws, 3, headers, rows)
        self._ajustar_larguras_por_headers(ws, headers)

    def _preencher_aba_levantamento(self, ws, projeto: ProjetoMemorial):
        ambientes = projeto.ambientes

        # ---- Título ----
        ws.merge_cells(f"B{self.ROW_TITULO}:M{self.ROW_TITULO}")
        c = ws.cell(
            row=self.ROW_TITULO,
            column=self.COL_NOME,
            value=f"MEMORIAL DE CÁLCULO — {projeto.nome_projeto.upper()}",
        )
        c.font = FONTE_TITULO
        c.fill = FILL_AZUL
        c.alignment = ALINHAR_CENTRO

        ws.merge_cells(f"B{self.ROW_SUBTITULO}:M{self.ROW_SUBTITULO}")
        c = ws.cell(
            row=self.ROW_SUBTITULO,
            column=self.COL_NOME,
            value="LEVANTAMENTO DE CAMPO — Alvenarias / Paredes",
        )
        c.font = Font(name="Arial", bold=True, size=11, color="FF003366")
        c.alignment = ALINHAR_CENTRO

        ws.row_dimensions[self.ROW_TITULO].height = 28
        ws.row_dimensions[self.ROW_SUBTITULO].height = 20

        # ---- Linha de grupo de colunas (ROW_HEADER1) ----
        grupos = [
            (self.COL_NOME, self.COL_SUBTITULO, "Identificação"),
            (self.COL_AREA, self.COL_PD, "Dimensões do Ambiente"),
            (self.COL_AREA_BRUTA, self.COL_AREA_LIQ, "Áreas de Parede"),
            (self.COL_ESP, self.COL_MATERIAL, "Composição"),
            (self.COL_PU, self.COL_TOTAL, "Custo (SINAPI-SP)"),
        ]
        for col_ini, col_fim, label in grupos:
            if col_ini == col_fim:
                ws.merge_cells(
                    start_row=self.ROW_HEADER1,
                    start_column=col_ini,
                    end_row=self.ROW_HEADER1,
                    end_column=col_fim,
                )
            else:
                ws.merge_cells(
                    start_row=self.ROW_HEADER1,
                    start_column=col_ini,
                    end_row=self.ROW_HEADER1,
                    end_column=col_fim,
                )
            c = ws.cell(row=self.ROW_HEADER1, column=col_ini, value=label)
            c.font = FONTE_HEADER
            c.fill = FILL_AZUL
            c.alignment = ALINHAR_CENTRO

        ws.row_dimensions[self.ROW_HEADER1].height = 18

        # ---- Cabeçalhos de coluna (ROW_HEADER2) ----
        headers = {
            self.COL_NOME: "Ambiente",
            self.COL_SUBTITULO: "Uso / Subtítulo",
            self.COL_AREA: "Área\n[m²]",
            self.COL_PERIMETRO: "Perímetro\n[m]",
            self.COL_PD: "Pé-direito\n[m]",
            self.COL_AREA_BRUTA: "Área bruta\nparede [m²]",
            self.COL_AREA_VAO: "Área de\nvãos [m²]",
            self.COL_AREA_LIQ: "Área líquida\nparede [m²]",
            self.COL_ESP: "Espessura\nparede [m]",
            self.COL_MATERIAL: "Material\nSINAPI",
            self.COL_PU: "Preço unit.\n[R$/m²]",
            self.COL_TOTAL: "Custo total\n[R$]",
        }
        for col, label in headers.items():
            c = ws.cell(row=self.ROW_HEADER2, column=col, value=label)
            c.font = Font(name="Arial", bold=True, size=9, color="FF000000")
            c.fill = FILL_CINZA
            c.alignment = ALINHAR_CENTRO
            c.border = BORDA_FINA

        ws.row_dimensions[self.ROW_HEADER2].height = 30

        # ---- Dados ----
        custo_total_geral = 0.0
        for idx, amb in enumerate(ambientes):
            row = self.ROW_DADOS_INI + idx
            fill = FILL_PAR if idx % 2 == 0 else None
            ws.row_dimensions[row].height = 15

            material = _categoria_sinapi(amb.nome)
            vals = {
                self.COL_NOME: amb.nome,
                self.COL_SUBTITULO: amb.subtitulo or "",
                self.COL_AREA: amb.area,
                self.COL_PERIMETRO: amb.perimetro or None,
                self.COL_PD: amb.pe_direito,
                self.COL_AREA_BRUTA: amb.area_bruta_parede,
                self.COL_AREA_VAO: amb.area_vaos,
                self.COL_AREA_LIQ: amb.area_liquida_parede,
                self.COL_ESP: amb.espessura_parede,
                self.COL_MATERIAL: material,
                self.COL_PU: amb.custo_unitario if amb.custo_unitario else None,
                self.COL_TOTAL: amb.custo_total if amb.custo_total else None,
            }
            formatos = {
                self.COL_AREA: FMT_NUMERO,
                self.COL_PERIMETRO: FMT_NUMERO,
                self.COL_PD: FMT_NUMERO,
                self.COL_AREA_BRUTA: FMT_NUMERO,
                self.COL_AREA_VAO: FMT_NUMERO,
                self.COL_AREA_LIQ: FMT_NUMERO,
                self.COL_ESP: FMT_NUMERO,
                self.COL_PU: FMT_MOEDA,
                self.COL_TOTAL: FMT_MOEDA,
            }

            for col, val in vals.items():
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONTE_DADO
                c.border = BORDA_FINA
                c.alignment = ALINHAR_CENTRO if col != self.COL_NOME else ALINHAR_ESQ
                if fill:
                    c.fill = fill
                if col in formatos and val is not None:
                    c.number_format = formatos[col]

            custo_total_geral += amb.custo_total or 0

        # ---- Linha de totais ----
        row_total = self.ROW_DADOS_INI + len(ambientes)
        ws.row_dimensions[row_total].height = 18

        ws.merge_cells(
            start_row=row_total,
            start_column=self.COL_NOME,
            end_row=row_total,
            end_column=self.COL_MATERIAL,
        )
        c = ws.cell(row=row_total, column=self.COL_NOME, value="TOTAL GERAL")
        c.font = FONTE_TOTAL
        c.fill = FILL_TOTAL
        c.alignment = ALINHAR_CENTRO
        c.border = BORDA_FINA

        # Soma área líquida
        r_ini = self.ROW_DADOS_INI
        r_fim = row_total - 1
        col_liq_l = get_column_letter(self.COL_AREA_LIQ)
        col_tot_l = get_column_letter(self.COL_TOTAL)

        # Custo total com fórmula SUM
        ct = ws.cell(
            row=row_total,
            column=self.COL_TOTAL,
            value=f"=SUM({col_tot_l}{r_ini}:{col_tot_l}{r_fim})",
        )
        ct.font = FONTE_TOTAL
        ct.fill = FILL_TOTAL
        ct.alignment = ALINHAR_DIR
        ct.border = BORDA_FINA
        ct.number_format = FMT_MOEDA

    @staticmethod
    def _ajustar_larguras(ws):
        larguras = {
            2: 26,  # B – Nome
            3: 18,  # C – Subtítulo
            4: 10,  # D – Área
            5: 11,  # E – Perímetro
            6: 10,  # F – PD
            7: 12,  # G – Área bruta
            8: 11,  # H – Área vão
            9: 13,  # I – Área líquida
            10: 11,  # J – Espessura
            11: 22,  # K – Material
            12: 14,  # L – PU
            13: 16,  # M – Total
        }
        for col, larg in larguras.items():
            ws.column_dimensions[get_column_letter(col)].width = larg
        # Ocultar coluna A
        ws.column_dimensions["A"].width = 2

    def generate(
        self,
        projeto: ProjetoMemorial,
        output_path: str,
        quantitativos: dict[str, Any] | None = None,
    ) -> Path:
        wb = self._criar_workbook(projeto, quantitativos)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

        out_p = Path(output_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)
        return out_p


# ---------------------------------------------------------------------------
# Função de integração (mantém assinatura original)
# ---------------------------------------------------------------------------
def _sincronizar_ambientes_com_quantitativos(ambientes, quantitativos: dict[str, Any] | None):
    if not ambientes or not quantitativos:
        return

    paredes = quantitativos.get("paredes", [])
    if not paredes:
        return

    total_comprimento = sum(p.get("comprimento_m") or 0 for p in paredes)
    total_area_bruta = sum(p.get("area_externa_m2") or 0 for p in paredes)
    total_volume_liquido = sum(p.get("volume_liquido_m3") or 0 for p in paredes)
    total_volume_bruto = sum(p.get("volume_bruto_m3") or 0 for p in paredes)
    total_descontos = sum(
        d.get("volume_descontado_m3") or 0
        for p in paredes
        for d in p.get("descontos_aberturas", [])
    )
    espessuras = [p.get("espessura_m") for p in paredes if p.get("espessura_m")]
    espessura = sum(espessuras) / len(espessuras) if espessuras else None

    if total_comprimento <= 0 or total_area_bruta <= 0:
        return

    pe_direito_medio = total_area_bruta / total_comprimento
    area_vaos = total_descontos / espessura if espessura else 0

    # Em DXFs simples sem textos de ambiente, o fallback cria um ambiente unico.
    # Aqui alinhamos essa linha com os quantitativos detalhados de parede.
    if len(ambientes) == 1 and ambientes[0].nome.startswith("AMBIENTE"):
        amb = ambientes[0]
        amb.perimetro = round(total_comprimento, 2)
        amb.pe_direito = round(pe_direito_medio, 2)
        if espessura:
            amb.espessura_parede = round(espessura, 3)
        amb.area_bruta_parede = round(total_area_bruta, 2)
        amb.area_vaos = round(area_vaos, 2)
        amb.area_liquida_parede = round(max(total_area_bruta - area_vaos, 0), 2)
        amb.comprimento_paredes = round(total_comprimento, 2)
        amb.comprimento_vaos = round(area_vaos / CADExtractor.ALTURA_VAO_PADRAO, 2)
        return

    if total_volume_bruto > 0 and total_volume_liquido > 0:
        fator_liquido = total_volume_liquido / total_volume_bruto
        for amb in ambientes:
            amb.area_vaos = round(amb.area_bruta_parede * (1 - fator_liquido), 2)
            amb.area_liquida_parede = round(
                max(amb.area_bruta_parede - amb.area_vaos, 0), 2
            )


def run_integration(
    dxf_file: str,
    template_file: str,
    output_file: str,
    quantitativos_dxf: dict[str, Any] | None = None,
):
    logger.info(f"Processando {os.path.basename(dxf_file)}...")

    basepath = Path(__file__).parent
    sinapi_path = basepath / "sinapi.xlsx"

    sinapi = carregar_sinapi(str(sinapi_path))
    logger.debug("carregamento do sinapi concluído")

    if quantitativos_dxf is None:
        try:
            from src.modules.drill import processar_dxf

            quantitativos_dxf = processar_dxf(dxf_file)
            if quantitativos_dxf.get("erro"):
                logger.warning(f"Quantitativos DXF indisponiveis: {quantitativos_dxf['erro']}")
                quantitativos_dxf = None
        except Exception as exc:
            logger.warning(f"Nao foi possivel extrair quantitativos detalhados do DXF: {exc}")

    extractor = CADExtractor(dxf_file)
    ambientes = extractor.extrair_dados_reais()
    if not ambientes:
        ambientes = _criar_ambientes_por_contornos(dxf_file, quantitativos_dxf)
        if ambientes:
            logger.info("Ambientes criados a partir de contornos fechados do DXF.")
    if not ambientes:
        ambientes = _criar_ambiente_por_quantitativos(quantitativos_dxf)
        if ambientes:
            logger.info("Ambiente tecnico criado a partir dos quantitativos do drill.py.")
    logger.info(f"Ambientes encontrados: {len(ambientes)}")
    if not ambientes:
        raise ValueError(
            "Nenhum ambiente foi encontrado no DXF. Verifique se a planta possui "
            "textos de ambiente/área ou contornos de parede em layers de parede/alvenaria."
        )
    _sincronizar_ambientes_com_quantitativos(ambientes, quantitativos_dxf)

    for a in ambientes:
        logger.info(
            f"  {a.nome:<30} área={a.area:>8.2f}m²  "
            f"P={a.perimetro:>7.2f}m  PD={a.pe_direito:.2f}m  "
            f"A_liq={a.area_liquida_parede:.2f}m²"
        )

    # Buscar preços SINAPI
    for ambiente in ambientes:
        material = _categoria_sinapi(ambiente.nome)
        preco = buscar_preco_sinapi(material, sinapi)

        if preco is not None:
            ambiente.custo_unitario = round(preco, 6)
            ambiente.custo_total = round(preco * (ambiente.area_liquida_parede or 0), 2)
            logger.info(
                f"[SINAPI] {ambiente.nome!r:<28} | {material!r:<25} | "
                f"R$ {ambiente.custo_unitario:.2f}/m² | total R$ {ambiente.custo_total:.2f}"
            )
        else:
            ambiente.custo_unitario = 0.0
            ambiente.custo_total = 0.0
            logger.warning(
                f"[SINAPI] Produto '{material}' não encontrado para '{ambiente.nome}'."
            )

    nome_projeto = (
        os.path.basename(dxf_file).replace(".dxf", "").replace(".DXF", "").title()
    )
    projeto = ProjetoMemorial(nome_projeto=nome_projeto, ambientes=ambientes)

    generator = MemorialGenerator(template_file)
    arquivo_final = generator.generate(projeto, output_file, quantitativos_dxf)

    logger.info(f"\nSucesso! Memorial gerado em: {arquivo_final}")
    return arquivo_final


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    basepath = Path(__file__).parent
    dxf_file = str((basepath / "teste.dxf").resolve())
    template_file = str((basepath / "model_memorial.xlsx").resolve())
    output_file = str((basepath / "memorial_preenchido.xlsx").resolve())

    run_integration(
        dxf_file=dxf_file, template_file=template_file, output_file=output_file
    )
