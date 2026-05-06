import openpyxl
from pathlib import Path
import json
from datetime import datetime

class MemorialPopulator:
    """
    Popula o arquivo Excel de memorial com os dados extraídos do DXF.
    Utiliza coordenadas de células baseadas na análise do modelo.
    """
    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path
        if not Path(template_path).exists():
            raise FileNotFoundError(f"Template não encontrado: {template_path}")
        self.wb = openpyxl.load_workbook(template_path)

    def populate_alvenaria(self, summary_data: dict):
        """
        Popula a planilha 'Alvenarias' com comprimentos de paredes.
        summary_data esperado: { "layer_name": { "LINE": {"total_length": 10.5}, ... } }
        """
        if 'Alvenarias' not in self.wb.sheetnames:
            print("Aviso: Planilha 'Alvenarias' não encontrada no template")
            return
            
        ws = self.wb['Alvenarias']
        row = 6
        total_geral = 0
        
        for layer, data in summary_data.items():
            if 'alv' in layer.lower() or 'parede' in layer.lower() or 'arq' in layer.lower():
                ws.cell(row=row, column=2).value = f"Layer: {layer}"
                total_len = 0
                for etype, metrics in data.items():
                    total_len += metrics.get("total_length", 0)
                ws.cell(row=row, column=6).value = round(total_len, 2)
                total_geral += total_len
                row += 1
        
        # Adiciona total geral
        if row > 6:
            ws.cell(row=row, column=2).value = "TOTAL GERAL"
            ws.cell(row=row, column=6).value = round(total_geral, 2)

    def populate_eletrica(self, synthesis_data: dict):
        """
        Popula a planilha 'Inst Elétricas'.
        synthesis_data: dados processados pelo Surveyor/Spatial Analyst
        """
        if 'Inst Elétricas' not in self.wb.sheetnames:
            print("Aviso: Planilha 'Inst Elétricas' não encontrada no template")
            return
            
        ws = self.wb['Inst Elétricas']
        row = 6
        total_geral = 0
        
        # Processo clusters se disponível
        if "clusters" in synthesis_data:
            for i, cluster in enumerate(synthesis_data["clusters"]):
                ws.cell(row=row + i, column=2).value = f"Circuito {i+1}"
                cluster_len = cluster.get("total_length", 0)
                ws.cell(row=row + i, column=6).value = cluster_len
                ws.cell(row=row + i, column=5).value = ", ".join(cluster.get("identifiers", []))
                total_geral += cluster_len
            row += len(synthesis_data["clusters"])
        
        # Adiciona total geral
        if row > 6:
            ws.cell(row=row + 1, column=2).value = "TOTAL GERAL"
            ws.cell(row=row + 1, column=6).value = round(total_geral, 2)

    def populate_hidraulica(self, synthesis_data: dict):
        """
        Popula a planilha 'Inst Hidráulicas' com dados de tubulações.
        """
        if 'Inst Hidráulicas' not in self.wb.sheetnames:
            print("Aviso: Planilha 'Inst Hidráulicas' não encontrada no template")
            return
            
        ws = self.wb['Inst Hidráulicas']
        row = 6
        total_geral = 0
        
        if "clusters" in synthesis_data:
            tipos_sistema = {"agua": "Água Fria", "esgoto": "Esgoto", "dreno": "Drenagem"}
            
            for i, cluster in enumerate(synthesis_data["clusters"]):
                identifiers = cluster.get("identifiers", [])
                
                # Determina o tipo de sistema
                sistema_tipo = "Tubulação"
                for key, label in tipos_sistema.items():
                    if any(key in str(id).lower() for id in identifiers):
                        sistema_tipo = label
                        break
                
                ws.cell(row=row + i, column=2).value = f"{sistema_tipo} {i+1}"
                cluster_len = cluster.get("total_length", 0)
                ws.cell(row=row + i, column=6).value = cluster_len
                ws.cell(row=row + i, column=5).value = ", ".join(identifiers)
                total_geral += cluster_len
            
            row += len(synthesis_data["clusters"])
        
        # Adiciona total geral
        if row > 6:
            ws.cell(row=row + 1, column=2).value = "TOTAL GERAL"
            ws.cell(row=row + 1, column=6).value = round(total_geral, 2)

    def populate_rede_dados(self, synthesis_data: dict):
        """
        Popula a planilha de Rede/Dados com comprimentos de cabos.
        """
        sheet_names = self.wb.sheetnames
        sheet_target = None
        
        for name in sheet_names:
            if 'rede' in name.lower() or 'dados' in name.lower() or 'comunicação' in name.lower():
                sheet_target = name
                break
        
        if not sheet_target:
            print("Aviso: Planilha de Rede/Dados não encontrada no template")
            return
        
        ws = self.wb[sheet_target]
        row = 6
        total_geral = 0
        
        if "clusters" in synthesis_data:
            for i, cluster in enumerate(synthesis_data["clusters"]):
                ws.cell(row=row + i, column=2).value = f"Cabo/Circuito {i+1}"
                cluster_len = cluster.get("total_length", 0)
                ws.cell(row=row + i, column=6).value = cluster_len
                ws.cell(row=row + i, column=5).value = ", ".join(cluster.get("identifiers", []))
                total_geral += cluster_len
            
            row += len(synthesis_data["clusters"])
        
        # Adiciona total geral
        if row > 6:
            ws.cell(row=row + 1, column=2).value = "TOTAL GERAL"
            ws.cell(row=row + 1, column=6).value = round(total_geral, 2)

    def add_metadata(self, disciplina: str, prompt: str = ""):
        """Adiciona metadados como disciplina, data e prompt na primeira planilha disponível."""
        ws = self.wb.active
        if ws:
            ws.cell(row=1, column=1).value = f"Disciplina: {disciplina.upper()}"
            ws.cell(row=2, column=1).value = f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            if prompt:
                ws.cell(row=3, column=1).value = f"Solicitação: {prompt[:60]}..."

    def save(self):
        self.wb.save(self.output_path)
        print(f"Memorial populado com sucesso em: {self.output_path}")

def run_population(data_json: str, discipline: str, template: str, output: str):
    """
    Função auxiliar para ser chamada como Tool pelos agentes.
    """
    try:
        data = json.loads(data_json)
        pop = MemorialPopulator(template, output)
        
        if "alvenaria" in discipline.lower():
            pop.populate_alvenaria(data)
        elif "eletrica" in discipline.lower() or "elétrica" in discipline.lower():
            pop.populate_eletrica(data)
        elif "hidraulica" in discipline.lower() or "hidráulica" in discipline.lower():
            pop.populate_hidraulica(data)
        elif "rede" in discipline.lower() or "dados" in discipline.lower():
            pop.populate_rede_dados(data)
        
        pop.add_metadata(discipline)
        pop.save()
        return f"Sucesso: Dados de {discipline} gravados no Excel."
    except Exception as e:
        return f"Erro na população: {str(e)}"
