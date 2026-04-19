
# generatorteste.py
# Gera o memorial de cálculo a partir dos dados extraídos do DXF.
# Cria um template novo e bem estruturado (não depende do model_memorial.xlsx
# para as colunas de dados), mas mantém a aba "Levantamento Campo" compatível.

import logging
import os
from pathlib import Path
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from dxf_extractor import Ambiente, CADExtractor, ProjetoMemorial
from sinapi import buscar_preco_sinapi, carregar_sinapi

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Mapeamento ambiente → material SINAPI
# ---------------------------------------------------------------------------
MAP_SINAPI = {
    'alojamento':   'bloco cerâmico',
    'banheiro':     'bloco cerâmico',
    'copa':         'bloco cerâmico',
    'sala':         'bloco cerâmico',
    'circulação':   'bloco cerâmico',
    'auditório':    'bloco cerâmico',
    'reserva':      'bloco cerâmico',
    'passadiço':    'bloco cerâmico',
    'calçada':      'concreto estrutural',
    'área':         'concreto estrutural',
    'telhado':      'telha cerâmica',
}

def _categoria_sinapi(nome: str) -> str:
    nome_l = nome.lower()
    for chave, material in MAP_SINAPI.items():
        if chave in nome_l:
            return material
    return 'bloco cerâmico'


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
COR_HEADER_AZUL   = 'FF003366'   # azul escuro Exército
COR_HEADER_CINZA  = 'FFD9D9D9'
COR_LINHA_PAR     = 'FFF5F5F5'
COR_TOTAL         = 'FFBDD7EE'
FONTE_TITULO      = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
FONTE_HEADER      = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
FONTE_SUBHEADER   = Font(name='Arial', bold=True, size=9,  color='FF000000')
FONTE_DADO        = Font(name='Arial', size=9)
FONTE_DADO_BOLD   = Font(name='Arial', bold=True, size=9)
FONTE_TOTAL       = Font(name='Arial', bold=True, size=9, color='FF003366')

FILL_AZUL   = PatternFill('solid', fgColor=COR_HEADER_AZUL)
FILL_CINZA  = PatternFill('solid', fgColor=COR_HEADER_CINZA)
FILL_PAR    = PatternFill('solid', fgColor=COR_LINHA_PAR)
FILL_TOTAL  = PatternFill('solid', fgColor=COR_TOTAL)

BORDA_FINA = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
BORDA_MEDIA = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium'),
)

ALINHAR_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALINHAR_ESQ    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALINHAR_DIR    = Alignment(horizontal='right',  vertical='center')

FMT_NUMERO   = '#,##0.00'
FMT_MOEDA    = 'R$ #,##0.00'
FMT_INTEIRO  = '#,##0'


def _aplicar_borda(cell, borda=BORDA_FINA):
    cell.border = borda


def _celula(ws, row, col, valor=None, fonte=None, fill=None, alinhamento=None,
            formato=None, borda=BORDA_FINA):
    c = ws.cell(row=row, column=col, value=valor)
    if fonte:        c.font = fonte
    if fill:         c.fill = fill
    if alinhamento:  c.alignment = alinhamento
    if formato:      c.number_format = formato
    if borda:        c.border = borda
    return c


# ---------------------------------------------------------------------------
# Gerador do memorial
# ---------------------------------------------------------------------------
class MemorialGenerator:

    # Colunas da aba "Levantamento Campo"
    # B  = Ambiente
    # C  = Subtítulo/uso
    # D  = Área (m²)
    # E  = Perímetro (m)
    # F  = Pé-direito (m)
    # G  = Área bruta de parede (m²)
    # H  = Área de vãos (m²)
    # I  = Área líquida de parede (m²)
    # J  = Espessura parede (m)
    # K  = Material / categoria
    # L  = Preço unitário SINAPI (R$/m²)
    # M  = Custo total (R$)

    COL_NOME       = 2   # B
    COL_SUBTITULO  = 3   # C
    COL_AREA       = 4   # D
    COL_PERIMETRO  = 5   # E
    COL_PD         = 6   # F
    COL_AREA_BRUTA = 7   # G
    COL_AREA_VAO   = 8   # H
    COL_AREA_LIQ   = 9   # I
    COL_ESP        = 10  # J
    COL_MATERIAL   = 11  # K
    COL_PU         = 12  # L
    COL_TOTAL      = 13  # M

    ROW_TITULO    = 1
    ROW_SUBTITULO = 2
    ROW_HEADER1   = 4
    ROW_HEADER2   = 5
    ROW_DADOS_INI = 6

    def __init__(self, template_path: str = None):
        # template_path é aceito mas ignorado — geramos nosso próprio template
        pass

    def _criar_workbook(self, projeto: ProjetoMemorial) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Levantamento Campo'

        self._preencher_aba_levantamento(ws, projeto)
        self._ajustar_larguras(ws)

        return wb

    def _preencher_aba_levantamento(self, ws, projeto: ProjetoMemorial):
        ambientes = projeto.ambientes

        # ---- Título ----
        ws.merge_cells(f'B{self.ROW_TITULO}:M{self.ROW_TITULO}')
        c = ws.cell(row=self.ROW_TITULO, column=self.COL_NOME,
                    value=f'MEMORIAL DE CÁLCULO — {projeto.nome_projeto.upper()}')
        c.font = FONTE_TITULO
        c.fill = FILL_AZUL
        c.alignment = ALINHAR_CENTRO

        ws.merge_cells(f'B{self.ROW_SUBTITULO}:M{self.ROW_SUBTITULO}')
        c = ws.cell(row=self.ROW_SUBTITULO, column=self.COL_NOME,
                    value='LEVANTAMENTO DE CAMPO — Alvenarias / Paredes')
        c.font = Font(name='Arial', bold=True, size=11, color='FF003366')
        c.alignment = ALINHAR_CENTRO

        ws.row_dimensions[self.ROW_TITULO].height    = 28
        ws.row_dimensions[self.ROW_SUBTITULO].height = 20

        # ---- Linha de grupo de colunas (ROW_HEADER1) ----
        grupos = [
            (self.COL_NOME,       self.COL_SUBTITULO, 'Identificação'),
            (self.COL_AREA,       self.COL_PD,        'Dimensões do Ambiente'),
            (self.COL_AREA_BRUTA, self.COL_AREA_LIQ,  'Áreas de Parede'),
            (self.COL_ESP,        self.COL_MATERIAL,   'Composição'),
            (self.COL_PU,         self.COL_TOTAL,      'Custo (SINAPI-SP)'),
        ]
        for col_ini, col_fim, label in grupos:
            if col_ini == col_fim:
                ws.merge_cells(
                    start_row=self.ROW_HEADER1, start_column=col_ini,
                    end_row=self.ROW_HEADER1,   end_column=col_fim)
            else:
                ws.merge_cells(
                    start_row=self.ROW_HEADER1, start_column=col_ini,
                    end_row=self.ROW_HEADER1,   end_column=col_fim)
            c = ws.cell(row=self.ROW_HEADER1, column=col_ini, value=label)
            c.font = FONTE_HEADER
            c.fill = FILL_AZUL
            c.alignment = ALINHAR_CENTRO

        ws.row_dimensions[self.ROW_HEADER1].height = 18

        # ---- Cabeçalhos de coluna (ROW_HEADER2) ----
        headers = {
            self.COL_NOME:       'Ambiente',
            self.COL_SUBTITULO:  'Uso / Subtítulo',
            self.COL_AREA:       'Área\n[m²]',
            self.COL_PERIMETRO:  'Perímetro\n[m]',
            self.COL_PD:         'Pé-direito\n[m]',
            self.COL_AREA_BRUTA: 'Área bruta\nparede [m²]',
            self.COL_AREA_VAO:   'Área de\nvãos [m²]',
            self.COL_AREA_LIQ:   'Área líquida\nparede [m²]',
            self.COL_ESP:        'Espessura\nparede [m]',
            self.COL_MATERIAL:   'Material\nSINAPI',
            self.COL_PU:         'Preço unit.\n[R$/m²]',
            self.COL_TOTAL:      'Custo total\n[R$]',
        }
        for col, label in headers.items():
            c = ws.cell(row=self.ROW_HEADER2, column=col, value=label)
            c.font = Font(name='Arial', bold=True, size=9, color='FF000000')
            c.fill = FILL_CINZA
            c.alignment = ALINHAR_CENTRO
            c.border = BORDA_FINA

        ws.row_dimensions[self.ROW_HEADER2].height = 30

        # ---- Dados ----
        custo_total_geral = 0.0
        for idx, amb in enumerate(ambientes):
            row = self.ROW_DADOS_INI + idx
            fill = FILL_PAR if idx % 2 == 0 else None
            ws.row_dimensions[row].height = 15

            material = _categoria_sinapi(amb.nome)
            vals = {
                self.COL_NOME:       amb.nome,
                self.COL_SUBTITULO:  amb.subtitulo or '',
                self.COL_AREA:       amb.area,
                self.COL_PERIMETRO:  amb.perimetro or None,
                self.COL_PD:         amb.pe_direito,
                self.COL_AREA_BRUTA: amb.area_bruta_parede,
                self.COL_AREA_VAO:   amb.area_vaos,
                self.COL_AREA_LIQ:   amb.area_liquida_parede,
                self.COL_ESP:        amb.espessura_parede,
                self.COL_MATERIAL:   material,
                self.COL_PU:         amb.custo_unitario if amb.custo_unitario else None,
                self.COL_TOTAL:      amb.custo_total if amb.custo_total else None,
            }
            formatos = {
                self.COL_AREA:       FMT_NUMERO,
                self.COL_PERIMETRO:  FMT_NUMERO,
                self.COL_PD:         FMT_NUMERO,
                self.COL_AREA_BRUTA: FMT_NUMERO,
                self.COL_AREA_VAO:   FMT_NUMERO,
                self.COL_AREA_LIQ:   FMT_NUMERO,
                self.COL_ESP:        FMT_NUMERO,
                self.COL_PU:         FMT_MOEDA,
                self.COL_TOTAL:      FMT_MOEDA,
            }

            for col, val in vals.items():
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONTE_DADO
                c.border = BORDA_FINA
                c.alignment = ALINHAR_CENTRO if col != self.COL_NOME else ALINHAR_ESQ
                if fill:
                    c.fill = fill
                if col in formatos and val is not None:
                    c.number_format = formatos[col]

            custo_total_geral += amb.custo_total or 0

        # ---- Linha de totais ----
        row_total = self.ROW_DADOS_INI + len(ambientes)
        ws.row_dimensions[row_total].height = 18

        ws.merge_cells(
            start_row=row_total, start_column=self.COL_NOME,
            end_row=row_total,   end_column=self.COL_MATERIAL)
        c = ws.cell(row=row_total, column=self.COL_NOME, value='TOTAL GERAL')
        c.font = FONTE_TOTAL
        c.fill = FILL_TOTAL
        c.alignment = ALINHAR_CENTRO
        c.border = BORDA_FINA

        # Soma área líquida
        r_ini = self.ROW_DADOS_INI
        r_fim = row_total - 1
        col_liq_l = get_column_letter(self.COL_AREA_LIQ)
        col_tot_l = get_column_letter(self.COL_TOTAL)

        # Custo total com fórmula SUM
        ct = ws.cell(row=row_total, column=self.COL_TOTAL,
                     value=f'=SUM({col_tot_l}{r_ini}:{col_tot_l}{r_fim})')
        ct.font = FONTE_TOTAL
        ct.fill = FILL_TOTAL
        ct.alignment = ALINHAR_DIR
        ct.border = BORDA_FINA
        ct.number_format = FMT_MOEDA

    @staticmethod
    def _ajustar_larguras(ws):
        larguras = {
            2: 26,   # B – Nome
            3: 18,   # C – Subtítulo
            4: 10,   # D – Área
            5: 11,   # E – Perímetro
            6: 10,   # F – PD
            7: 12,   # G – Área bruta
            8: 11,   # H – Área vão
            9: 13,   # I – Área líquida
            10: 11,  # J – Espessura
            11: 22,  # K – Material
            12: 14,  # L – PU
            13: 16,  # M – Total
        }
        for col, larg in larguras.items():
            ws.column_dimensions[get_column_letter(col)].width = larg
        # Ocultar coluna A
        ws.column_dimensions['A'].width = 2

    def generate(self, projeto: ProjetoMemorial, output_path: str) -> Path:
        wb = self._criar_workbook(projeto)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

        out_p = Path(output_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)
        return out_p


# ---------------------------------------------------------------------------
# Função de integração (mantém assinatura original)
# ---------------------------------------------------------------------------
def run_integration(dxf_file: str, template_file: str, output_file: str):
    logger.info(f"Processando {os.path.basename(dxf_file)}...")

    basepath = Path(__file__).parent
    sinapi_path = basepath / 'sinapi.xlsx'

    sinapi = carregar_sinapi(str(sinapi_path))

    extractor = CADExtractor(dxf_file)
    ambientes = extractor.extrair_dados_reais()
    logger.info(f"Ambientes encontrados: {len(ambientes)}")

    for a in ambientes:
        logger.info(
            f"  {a.nome:<30} área={a.area:>8.2f}m²  "
            f"P={a.perimetro:>7.2f}m  PD={a.pe_direito:.2f}m  "
            f"A_liq={a.area_liquida_parede:.2f}m²"
        )

    # Buscar preços SINAPI
    for ambiente in ambientes:
        material = _categoria_sinapi(ambiente.nome)
        preco = buscar_preco_sinapi(material, sinapi)

        if preco is not None:
            ambiente.custo_unitario = round(preco, 6)
            ambiente.custo_total = round(preco * (ambiente.area_liquida_parede or 0), 2)
            logger.info(
                f"[SINAPI] {ambiente.nome!r:<28} | {material!r:<25} | "
                f"R$ {ambiente.custo_unitario:.2f}/m² | total R$ {ambiente.custo_total:.2f}"
            )
        else:
            ambiente.custo_unitario = 0.0
            ambiente.custo_total    = 0.0
            logger.warning(f"[SINAPI] Produto '{material}' não encontrado para '{ambiente.nome}'.")

    nome_projeto = os.path.basename(dxf_file).replace('.dxf', '').replace('.DXF', '').title()
    projeto = ProjetoMemorial(nome_projeto=nome_projeto, ambientes=ambientes)

    generator = MemorialGenerator(template_file)
    arquivo_final = generator.generate(projeto, output_file)

    logger.info(f"\nSucesso! Memorial gerado em: {arquivo_final}")
    return arquivo_final


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    basepath = Path(__file__).parent
    dxf_file      = str((basepath / 'teste.dxf').resolve())
    template_file = str((basepath / 'model_memorial.xlsx').resolve())
    output_file   = str((basepath / 'memorial_preenchido.xlsx').resolve())

    run_integration(dxf_file=dxf_file, template_file=template_file, output_file=output_file)